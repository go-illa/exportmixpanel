import os
import io
import requests
import openpyxl
from openpyxl import Workbook
from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    redirect,
    url_for,
    flash,
    send_file,
    session as flask_session
)
from sqlalchemy import create_engine
from sqlalchemy.orm import scoped_session, sessionmaker
from datetime import datetime, timedelta
import shutil
import subprocess
from collections import defaultdict, Counter
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
import math
import hashlib
import json
import concurrent.futures
import pandas as pd
import traceback
import logging
import re
import time
from threading import Thread

from db.config import DB_URI, API_TOKEN, BASE_API_URL, API_EMAIL, API_PASSWORD
from db.models import Base, Trip, Tag

# Import the export_data_for_comparison function
from exportmix import export_data_for_comparison

app = Flask(__name__)
engine = create_engine(
    DB_URI,
    pool_size=20,         # Increase the default pool size
    max_overflow=20,      # Allow more connections to overflow
    pool_timeout=30       # How long to wait for a connection to become available
    )
db_session = scoped_session(sessionmaker(autocommit=False, autoflush=False, bind=engine))
update_jobs = {}
executor = ThreadPoolExecutor(max_workers=40)
app.secret_key = "your_secret_key"  # for flashing and session

# Global dict to track progress of long-running operations
progress_data = {}

# Helper function for calculating haversine distance between two coordinates
def haversine_distance(coord1, coord2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    
    Args:
        coord1: tuple or list with (lat, lon)
        coord2: tuple or list with (lat, lon)
        
    Returns:
        Distance in kilometers
    """
    lat1, lon1 = coord1
    lat2, lon2 = coord2
    
    # Convert decimal degrees to radians
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    
    # Haversine formula
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    r = 6371  # Radius of earth in kilometers
    return c * r

def calculate_expected_trip_quality(
    logs_count, 
    lack_of_accuracy, 
    medium_segments_count, 
    long_segments_count, 
    short_dist_total, 
    medium_dist_total, 
    long_dist_total,
    calculated_distance
):
    """
    Enhanced expected trip quality calculation.
    
    Special cases:
      - "No Logs Trip": if logs_count <= 1 OR if calculated_distance <= 0 OR if the total recorded distance 
         (short_dist_total + medium_dist_total + long_dist_total) is <= 0.
      - "Trip Points Only Exist": if logs_count < 50 but there is at least one medium or long segment.
    
    Otherwise, the quality score is calculated as follows:
    
      1. Normalize the logs count:
         LF = min(logs_count / 500, 1)
         
      2. Compute the ratio of short-distance to (medium + long) distances:
         R = short_dist_total / (medium_dist_total + long_dist_total + ε)
         
      3. Determine the segment factor SF:
         SF = 1 if R ≥ 5,
              = 0 if R ≤ 0.5,
              = (R - 0.5) / 4.5 otherwise.
         
      4. Compute the overall quality score:
         Q = 0.5 × LF + 0.5 × SF
         
      5. If lack_of_accuracy is True, penalize Q by 20% (i.e. Q = 0.8 × Q).
         
      6. Map Q to a quality category:
         - Q ≥ 0.8: "High Quality Trip"
         - 0.5 ≤ Q < 0.8: "Moderate Quality Trip"
         - Q < 0.5: "Low Quality Trip"
    
    Returns:
      str: Expected trip quality category.
    """
    epsilon = 1e-2  # Small constant to avoid division by zero

    # NEW: If the calculated distance is zero (or non-positive) OR if there is essentially no recorded distance,
    # return "No Logs Trip"
    if (short_dist_total + medium_dist_total + long_dist_total) <= 0 or logs_count <= 1:
        return "No Logs Trip"

    # Special condition: very few logs and no medium or long segments.
    if logs_count < 5 and medium_segments_count == 0 and long_segments_count == 0:
        return "No Logs Trip"

    # Special condition: few logs (<50) but with some medium or long segments.
    if logs_count < 50 and (medium_segments_count > 0 or long_segments_count > 0):
        return "Trip Points Only Exist"
    
    # 1. Normalize the logs count (saturate at 500)
    logs_factor = min(logs_count / 500.0, 1.0)
    
    # 2. Compute the ratio of short to (medium + long) distances
    ratio = short_dist_total / (medium_dist_total + long_dist_total + epsilon)
    
    # 3. Compute the segment factor based on ratio R
    if ratio >= 5:
        segment_factor = 1.0
    elif ratio <= 0.5:
        segment_factor = 0.0
    else:
        segment_factor = (ratio - 0.5) / 4.5
    
    # 4. Compute the overall quality score Q
    quality_score = 0.5 * logs_factor + 0.5 * segment_factor
    
    # 5. Apply penalty if GPS accuracy is lacking
    if lack_of_accuracy:
        quality_score *= 0.8

    # 6. Map the quality score to a quality category
    if quality_score >= 0.8 and (medium_dist_total + long_dist_total) <= 0.05*calculated_distance:
        return "High Quality Trip"
    elif quality_score >= 0.5:
        return "Moderate Quality Trip"
    else:
        return "Low Quality Trip"







# Function to analyze trip segments and distances
def analyze_trip_segments(coordinates):
    """
    Analyze coordinates to calculate distance metrics:
    - Count and total distance of short segments (<1km)
    - Count and total distance of medium segments (1-5km)
    - Count and total distance of long segments (>5km)
    - Maximum segment distance
    - Average segment distance
    
    Args:
        coordinates: list of [lon, lat] points from API
        
    Returns:
        Dictionary with analysis metrics
    """
    if not coordinates or len(coordinates) < 2:
        return {
            "short_segments_count": 0,
            "medium_segments_count": 0,
            "long_segments_count": 0,
            "short_segments_distance": 0,
            "medium_segments_distance": 0,
            "long_segments_distance": 0,
            "max_segment_distance": 0,
            "avg_segment_distance": 0
        }
    
    # Note: API returns coordinates as [lon, lat], so we need to swap
    # Let's convert to [lat, lon] for calculations
    coords = [[float(point[1]), float(point[0])] for point in coordinates]
    
    short_segments_count = 0
    medium_segments_count = 0
    long_segments_count = 0
    short_segments_distance = 0
    medium_segments_distance = 0
    long_segments_distance = 0
    max_segment_distance = 0
    total_distance = 0
    segment_count = 0
    
    for i in range(len(coords) - 1):
        distance = haversine_distance(coords[i], coords[i+1])
        segment_count += 1
        total_distance += distance
        
        if distance < 1:
            short_segments_count += 1
            short_segments_distance += distance
        elif distance <= 5:
            medium_segments_count += 1
            medium_segments_distance += distance
        else:
            long_segments_count += 1
            long_segments_distance += distance
            
        if distance > max_segment_distance:
            max_segment_distance = distance
            
    avg_segment_distance = total_distance / segment_count if segment_count > 0 else 0
    
    return {
        "short_segments_count": short_segments_count,
        "medium_segments_count": medium_segments_count,
        "long_segments_count": long_segments_count,
        "short_segments_distance": round(short_segments_distance, 2),
        "medium_segments_distance": round(medium_segments_distance, 2),
        "long_segments_distance": round(long_segments_distance, 2),
        "max_segment_distance": round(max_segment_distance, 2),
        "avg_segment_distance": round(avg_segment_distance, 2)
    }


# --- Begin Migration to update schema with new columns ---
def migrate_db():
    try:
        print("Creating database tables from models...")
        Base.metadata.create_all(bind=engine)
        print("Database tables created successfully")
    except Exception as e:
        app.logger.error(f"Migration error: {e}")
        print(f"Error during database migration: {e}")

print("Running database migration...")
migrate_db()
print("Database migration completed")
# --- End Migration ---

@app.teardown_appcontext
def shutdown_session(exception=None):
    db_session.remove()

# ---------------------------
# Utility Functions
# ---------------------------

def get_saved_filters():
    return flask_session.get("saved_filters", {})

def save_filter_to_session(name, filters):
    saved = flask_session.get("saved_filters", {})
    saved[name] = filters
    flask_session["saved_filters"] = saved

def fetch_api_token():
    url = f"{BASE_API_URL}/auth/sign_in"
    payload = {"admin_user": {"email": API_EMAIL, "password": API_PASSWORD}}
    resp = requests.post(url, json=payload)
    if resp.status_code == 200:
        return resp.json().get("token", None)
    else:
        print("Error fetching primary token:", resp.text)
        return None

def fetch_api_token_alternative():
    alt_email = "SupplyPartner@illa.com.eg"
    alt_password = "654321"
    url = f"{BASE_API_URL}/auth/sign_in"
    payload = {"admin_user": {"email": alt_email, "password": alt_password}}
    try:
        resp = requests.post(url, json=payload)
        resp.raise_for_status()
        return resp.json().get("token", None)
    except Exception as e:
        print("Error fetching alternative token:", e)
        return None

def load_excel_data(excel_path):
    if not os.path.exists(excel_path):
        print(f"Excel file not found: {excel_path}. Returning empty data.")
        return []
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return []
    
    sheet = workbook.active
    headers = []
    data = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            headers = row
        else:
            row_dict = {headers[j]: row[j] for j in range(len(row))}
            data.append(row_dict)
    print(f"Loaded {len(data)} rows from Excel.")
    return data


# Carrier grouping
CARRIER_GROUPS = {
    "Vodafone": ["vodafone", "voda fone", "tegi ne3eesh"],
    "Orange": ["orange", "orangeeg", "orange eg"],
    "Etisalat": ["etisalat", "e& etisalat", "e&"],
    "We": ["we"]
}

def normalize_carrier(carrier_name):
    if not carrier_name:
        return ""
    lower = carrier_name.lower().strip()
    for group, variants in CARRIER_GROUPS.items():
        for variant in variants:
            if variant in lower:
                return group
    return carrier_name.title()

# NEW FUNCTION: determine_completed_by
# This function inspects an activity list to find the latest event where the status changes to 'completed'
# and returns the corresponding user_type (admin or driver), or None if not found.
def determine_completed_by(activity_list):
    best_candidate = None
    best_time = None
    for event in activity_list:
        changes = event.get("changes", {})
        status_change = changes.get("status")
        if status_change and isinstance(status_change, list) and len(status_change) >= 2:
            if str(status_change[-1]).lower() == "completed":
                created_str = event.get("created_at", "").replace(" UTC", "")
                event_time = None
                for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%fZ"]:
                    try:
                        event_time = datetime.strptime(created_str, fmt)
                        break
                    except ValueError:
                        continue
                if event_time:
                    if best_time is None or event_time > best_time:
                        best_time = event_time
                        best_candidate = event
    if best_candidate:
        return best_candidate.get("user_type", None)
    return None

# This function calculates the trip time (in hours) based on the time difference
# between the first arrival event and the completion event from the activity list
def calculate_trip_time(activity_list):
    arrival_time = None
    completion_time = None
    
    # Find first arrival time (status changes from pending to arrived)
    for event in activity_list:
        changes = event.get("changes", {})
        status_change = changes.get("status")
        if status_change and isinstance(status_change, list) and len(status_change) >= 2:
            if str(status_change[0]).lower() == "pending" and str(status_change[1]).lower() == "arrived":
                created_str = event.get("created_at", "").replace(" UTC", "")
                for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%fZ"]:
                    try:
                        arrival_time = datetime.strptime(created_str, fmt)
                        break
                    except ValueError:
                        continue
                if arrival_time:
                    break  # Found the first arrival time, so stop looking
    
    # Find completion time (status changes to completed)
    for event in activity_list:
        changes = event.get("changes", {})
        status_change = changes.get("status")
        if status_change and isinstance(status_change, list) and len(status_change) >= 2:
            if str(status_change[1]).lower() == "completed":
                created_str = event.get("created_at", "").replace(" UTC", "")
                for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%fZ"]:
                    try:
                        completion_time = datetime.strptime(created_str, fmt)
                        break
                    except ValueError:
                        continue
    
    # Calculate trip time in hours if both times were found
    if arrival_time and completion_time:
        time_diff = completion_time - arrival_time
        hours = time_diff.total_seconds() / 3600.0
        return round(hours, 2)  # Round to 2 decimal places
    
    return None

def fetch_coordinates_count(trip_id, token=API_TOKEN):
    url = f"{BASE_API_URL}/trips/{trip_id}/coordinates"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        data = resp.json()
        # Return the 'count' from the attributes; default to 0 if not found
        return data["data"]["attributes"].get("count", 0)
    except Exception as e:
        print(f"Error fetching coordinates for trip {trip_id}: {e}")
        return None

def fetch_trip_from_api(trip_id, token=API_TOKEN):
    url = f"{BASE_API_URL}/trips/{trip_id}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        data = resp.json()
        calc = data.get("data", {}).get("attributes", {}).get("calculatedDistance")
        if not calc or calc in [None, "", "N/A"]:
            raise ValueError("Missing calculatedDistance")
        return data
    except Exception as e:
        print("Error fetching trip data with primary token:", e)
        alt_token = fetch_api_token_alternative()
        if alt_token:
            headers = {"Authorization": f"Bearer {alt_token}", "Content-Type": "application/json"}
            try:
                resp = requests.get(url, headers=headers)
                resp.raise_for_status()
                data = resp.json()
                data["used_alternative"] = True
                return data
            except requests.HTTPError as http_err:
                if resp.status_code == 404:
                    print(f"Trip {trip_id} not found with alternative token (404).")
                else:
                    print(f"HTTP error with alternative token for trip {trip_id}: {http_err}")
            except Exception as e:
                print(f"Alternative fetch failed for trip {trip_id}: {e}")
        else:
            return None

def update_trip_db(trip_id, force_update=False, session_local=None):
    """
    Update or create trip record in database
    
    Args:
        trip_id: The trip ID to update
        force_update: If True, fetch from API even if record exists
        session_local: Optional db session to use
        
    Returns:
        Tuple of (Trip object, update status dict)
    """
    close_session = False
    if session_local is None:
        session_local = db_session()
        close_session = True
    
    # Flags to ensure alternative is only tried once
    tried_alternative_for_main = False
    tried_alternative_for_coordinate = False
    
    # Track what was updated for better reporting
    update_status = {
        "needed_update": False,
        "record_exists": False,
        "updated_fields": [],
        "reason_for_update": []
    }

    try:
        # Check if trip exists in database
        db_trip = session_local.query(Trip).filter(Trip.trip_id == trip_id).first()
        
        # If trip exists and data is complete and force_update is False, return it without API call
        if db_trip and not force_update and _is_trip_data_complete(db_trip):
            app.logger.debug(f"Trip {trip_id} already has complete data, skipping API call")
            return db_trip, update_status
        
        # Helper to validate field values
        def is_valid(value):
            return value is not None and str(value).strip() != "" and str(value).strip().upper() != "N/A"
        

        # Step 1: Check if trip exists and what fields need updating
        if db_trip:
            update_status["record_exists"] = True
            
            # If we're forcing an update, don't bother checking what's missing
            if force_update:
                update_status["needed_update"] = True
                update_status["reason_for_update"].append("Forced update")
            else:
                # Otherwise, check each field to see what needs updating
                missing_fields = []
                
                # Check manual_distance
                if not is_valid(db_trip.manual_distance):
                    missing_fields.append("manual_distance")
                    update_status["reason_for_update"].append("Missing manual_distance")
                
                # Check calculated_distance
                if not is_valid(db_trip.calculated_distance):
                    missing_fields.append("calculated_distance")
                    update_status["reason_for_update"].append("Missing calculated_distance")
                
                # Check trip_time
                if not is_valid(db_trip.trip_time):
                    missing_fields.append("trip_time")
                    update_status["reason_for_update"].append("Missing trip_time")
                
                # Check completed_by
                if not is_valid(db_trip.completed_by):
                    missing_fields.append("completed_by")
                    update_status["reason_for_update"].append("Missing completed_by")
                
                # Check coordinate_count
                if not is_valid(db_trip.coordinate_count):
                    missing_fields.append("coordinate_count")
                    update_status["reason_for_update"].append("Missing coordinate_count")
                
                # Check lack_of_accuracy (boolean should be explicitly set)
                if db_trip.lack_of_accuracy is None:
                    missing_fields.append("lack_of_accuracy")
                    update_status["reason_for_update"].append("Missing lack_of_accuracy")
                
                # Check segment counts
                if not is_valid(db_trip.short_segments_count):
                    missing_fields.append("segment_counts")
                    update_status["reason_for_update"].append("Missing segment counts")
                elif not is_valid(db_trip.medium_segments_count):
                    missing_fields.append("segment_counts")
                    update_status["reason_for_update"].append("Missing segment counts")
                elif not is_valid(db_trip.long_segments_count):
                    missing_fields.append("segment_counts")
                    update_status["reason_for_update"].append("Missing segment counts")
                
                # If no missing fields, return the trip without further API calls
                if not missing_fields:
                    return db_trip, update_status
                
                # Mark that this record needs update
                update_status["needed_update"] = True
        else:
            # Trip doesn't exist, so we'll create it
            update_status["needed_update"] = True
            update_status["reason_for_update"].append("New record")
            # Create an empty trip record that we'll populate later
            db_trip = Trip(trip_id=trip_id)
            session_local.add(db_trip)
            # Add all fields to missing_fields to ensure we fetch everything
            missing_fields = ["manual_distance", "calculated_distance", "trip_time", 
                             "completed_by", "coordinate_count", "lack_of_accuracy", 
                             "segment_counts"]
        
        # Step 2: Only proceed with API calls if the trip needs updating
        if update_status["needed_update"] or force_update:
            
            # Determine what API calls we need to make based on missing fields
            need_main_data = force_update or any(field in missing_fields for field 
                                                 in ["manual_distance", "calculated_distance", 
                                                     "trip_time", "completed_by", "lack_of_accuracy"])
            
            need_coordinates = force_update or "coordinate_count" in missing_fields
            
            need_segments = force_update or "segment_counts" in missing_fields
            
            # Step 2a: Fetch main trip data if needed
            if need_main_data:
                api_data = fetch_trip_from_api(trip_id)
                
                # If initial fetch fails, try alternative token
                if not (api_data and "data" in api_data):
                    if not tried_alternative_for_main:
                        tried_alternative_for_main = True
                        alt_token = fetch_api_token_alternative()
                        if alt_token:
                            headers = {"Authorization": f"Bearer {alt_token}", "Content-Type": "application/json"}
                            url = f"{BASE_API_URL}/trips/{trip_id}"
                            try:
                                resp = requests.get(url, headers=headers)
                                resp.raise_for_status()
                                api_data = resp.json()
                                api_data["used_alternative"] = True
                            except requests.HTTPError as http_err:
                                if resp.status_code == 404:
                                    print(f"Trip {trip_id} not found with alternative token (404).")
                                else:
                                    print(f"HTTP error with alternative token for trip {trip_id}: {http_err}")
                            except Exception as e:
                                print(f"Alternative fetch failed for trip {trip_id}: {e}")
                
                # Process the trip data if we got it
                if api_data and "data" in api_data:
                    trip_attributes = api_data["data"]["attributes"]
                    
                    # Update status regardless of what fields need updating
                    old_status = db_trip.status
                    db_trip.status = trip_attributes.get("status")
                    if db_trip.status != old_status:
                        update_status["updated_fields"].append("status")
                    
                    # Update manual_distance if needed
                    if force_update or "manual_distance" in missing_fields:
                        try:
                            old_value = db_trip.manual_distance
                            db_trip.manual_distance = float(trip_attributes.get("manualDistance") or 0)
                            if db_trip.manual_distance != old_value:
                                update_status["updated_fields"].append("manual_distance")
                        except ValueError:
                            db_trip.manual_distance = None
                    
                    # Update calculated_distance if needed
                    if force_update or "calculated_distance" in missing_fields:
                        try:
                            old_value = db_trip.calculated_distance
                            db_trip.calculated_distance = float(trip_attributes.get("calculatedDistance") or 0)
                            if db_trip.calculated_distance != old_value:
                                update_status["updated_fields"].append("calculated_distance")
                        except ValueError:
                            db_trip.calculated_distance = None
                    
                    # Mark supply partner if needed
                    if api_data.get("used_alternative"):
                        db_trip.supply_partner = True
                    
                    # Process trip_time only if missing or force_update
                    if force_update or "trip_time" in missing_fields:
                        activity_list = trip_attributes.get("activity", [])
                        trip_time = calculate_trip_time(activity_list)
                        
                        if trip_time is not None:
                            old_value = db_trip.trip_time
                            db_trip.trip_time = trip_time
                            if db_trip.trip_time != old_value:
                                update_status["updated_fields"].append("trip_time")
                                app.logger.info(f"Trip {trip_id}: trip_time updated to {trip_time} hours based on activity events")
                    
                    # Determine completed_by if missing or force_update
                    if force_update or "completed_by" in missing_fields:
                        comp_by = determine_completed_by(trip_attributes.get("activity", []))
                        if comp_by is not None:
                            old_value = db_trip.completed_by
                            db_trip.completed_by = comp_by
                            if db_trip.completed_by != old_value:
                                update_status["updated_fields"].append("completed_by")
                            app.logger.info(f"Trip {trip_id}: completed_by set to {db_trip.completed_by} based on activity events")
                        else:
                            db_trip.completed_by = None
                            app.logger.info(f"Trip {trip_id}: No completion event found, completed_by remains None")
                    
                    # Update lack_of_accuracy if missing or force_update
                    if force_update or "lack_of_accuracy" in missing_fields:
                        old_value = db_trip.lack_of_accuracy
                        tags_count = api_data["data"]["attributes"].get("tagsCount", [])
                        if isinstance(tags_count, list) and any(item.get("tag_name") == "lack_of_accuracy" and int(item.get("count", 0)) > 0 for item in tags_count):
                            db_trip.lack_of_accuracy = True
                        else:
                            db_trip.lack_of_accuracy = False
                        if db_trip.lack_of_accuracy != old_value:
                            update_status["updated_fields"].append("lack_of_accuracy")
            
            # Step 2b: Fetch coordinate count if needed
            if need_coordinates:
                coordinate_count = fetch_coordinates_count(trip_id)
                
                # Try alternative token if needed
                if not is_valid(coordinate_count) and not tried_alternative_for_coordinate:
                    tried_alternative_for_coordinate = True
                    alt_token = fetch_api_token_alternative()
                    if alt_token:
                        coordinate_count = fetch_coordinates_count(trip_id, token=alt_token)
                
                # Update the coordinate count if it changed
                if coordinate_count != db_trip.coordinate_count:
                    db_trip.coordinate_count = coordinate_count
                    update_status["updated_fields"].append("coordinate_count")
            
            # Step 2c: Fetch segment analysis if needed
            if need_segments:
                # Fetch coordinates
                url = f"{BASE_API_URL}/trips/{trip_id}/coordinates"
                token = fetch_api_token() or API_TOKEN
                headers = {
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/json"
                }
                
                try:
                    resp = requests.get(url, headers=headers)
                    # If unauthorized, try alternative token
                    if resp.status_code == 401:
                        alt_token = fetch_api_token_alternative()
                        if alt_token:
                            headers["Authorization"] = f"Bearer {alt_token}"
                            resp = requests.get(url, headers=headers)
                    
                    resp.raise_for_status()
                    coordinates_data = resp.json()
                    
                    if coordinates_data and "data" in coordinates_data and "attributes" in coordinates_data["data"]:
                        coordinates = coordinates_data["data"]["attributes"].get("coordinates", [])
                        
                        if coordinates and len(coordinates) >= 2:
                            analysis = analyze_trip_segments(coordinates)
                            
                            # Check if any segment metrics have changed
                            segments_changed = False
                            for key, value in analysis.items():
                                if getattr(db_trip, key, None) != value:
                                    segments_changed = True
                                    break
                                    
                            # Update trip with analysis results
                            db_trip.short_segments_count = analysis["short_segments_count"]
                            db_trip.medium_segments_count = analysis["medium_segments_count"]
                            db_trip.long_segments_count = analysis["long_segments_count"]
                            db_trip.short_segments_distance = analysis["short_segments_distance"]
                            db_trip.medium_segments_distance = analysis["medium_segments_distance"]
                            db_trip.long_segments_distance = analysis["long_segments_distance"]
                            db_trip.max_segment_distance = analysis["max_segment_distance"]
                            db_trip.avg_segment_distance = analysis["avg_segment_distance"]
                            
                            if segments_changed:
                                update_status["updated_fields"].append("segment_metrics")
                                
                            app.logger.info(f"Trip {trip_id}: Updated distance analysis metrics")
                        else:
                            app.logger.info(f"Trip {trip_id}: Not enough coordinates for detailed analysis")
                    
                    # Regardless of whether enough coordinates were fetched,
                    # always compute Expected Trip Quality using current DB values.
                    expected_quality = calculate_expected_trip_quality(
                        logs_count = db_trip.coordinate_count if db_trip.coordinate_count is not None else 0,
                        lack_of_accuracy = db_trip.lack_of_accuracy if db_trip.lack_of_accuracy is not None else False,
                        medium_segments_count = db_trip.medium_segments_count if db_trip.medium_segments_count is not None else 0,
                        long_segments_count = db_trip.long_segments_count if db_trip.long_segments_count is not None else 0,
                        short_dist_total = db_trip.short_segments_distance if db_trip.short_segments_distance is not None else 0.0,
                        medium_dist_total = db_trip.medium_segments_distance if db_trip.medium_segments_distance is not None else 0.0,
                        long_dist_total = db_trip.long_segments_distance if db_trip.long_segments_distance is not None else 0.0,
                        calculated_distance = db_trip.calculated_distance if db_trip.calculated_distance is not None else 0.0
                    )
                    if db_trip.expected_trip_quality != expected_quality:
                        db_trip.expected_trip_quality = expected_quality
                        update_status["updated_fields"].append("expected_trip_quality")
                    app.logger.info(f"Trip {trip_id}: Expected Trip Quality updated to '{expected_quality}'")
                    
                except Exception as e:
                    app.logger.error(f"Error fetching coordinates for trip {trip_id}: {e}")
            
            # If we made any updates, commit them
            if update_status["updated_fields"]:
                session_local.commit()
                session_local.refresh(db_trip)
            
        return db_trip, update_status
    except Exception as e:
        print("Error in update_trip_db:", e)
        session_local.rollback()
        db_trip = session_local.query(Trip).filter_by(trip_id=trip_id).first()
        return db_trip, {"error": str(e)}
    finally:
        if close_session:
            session_local.close()


# ---------------------------
# Routes 
# ---------------------------

@app.route("/update_db", methods=["POST"])
def update_db():
    """
    Bulk update DB from Excel (fetch each trip from the API) with improved performance.
    Only fetches data for trips that are missing critical fields or where force_update is True.
    Uses threading for faster processing.
    """
    import concurrent.futures
    
    session_local = db_session()
    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)
    
    # Track statistics
    stats = {
        "total": 0,
        "updated": 0,
        "skipped": 0,
        "errors": 0,
        "created": 0,
        "updated_fields": Counter(),  # Count which fields were updated most often
        "reasons": Counter()          # Count reasons for updates
    }
    
    # Get all trip IDs from Excel
    trip_ids = [row.get("tripId") for row in excel_data if row.get("tripId")]
    stats["total"] = len(trip_ids)
    
    # Define a worker function for thread pool
    def process_trip(trip_id):
        trip_stats = {
            "updated": 0,
            "skipped": 0,
            "errors": 0,
            "created": 0,
            "updated_fields": Counter(),
            "reasons": Counter()
        }
        
        # Create a new session for each thread to avoid conflicts
        thread_session = db_session()
        
        try:
            # False means don't force updates if all fields are present
            db_trip, update_status = update_trip_db(trip_id, force_update=False, session_local=thread_session)
            
            # Track statistics
            if "error" in update_status:
                trip_stats["errors"] += 1
            elif not update_status["record_exists"]:
                trip_stats["created"] += 1
                trip_stats["updated"] += 1
                # Count which fields were updated
                for field in update_status["updated_fields"]:
                    trip_stats["updated_fields"][field] += 1
            elif update_status["updated_fields"]:
                trip_stats["updated"] += 1
                # Count which fields were updated
                for field in update_status["updated_fields"]:
                    trip_stats["updated_fields"][field] += 1
            else:
                trip_stats["skipped"] += 1
                
            # Track reasons for updates
            for reason in update_status["reason_for_update"]:
                trip_stats["reasons"][reason] += 1
                
        except Exception as e:
            trip_stats["errors"] += 1
            print(f"Error processing trip {trip_id}: {e}")
        finally:
            thread_session.close()
            
        return trip_stats
    
    # Use ThreadPoolExecutor to process trips in parallel
    # Number of workers should be adjusted based on system capability and API rate limits
    max_workers = min(32, (os.cpu_count() or 1) * 4)  # Adjust based on system capability
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all trips to the executor
        future_to_trip = {executor.submit(process_trip, trip_id): trip_id for trip_id in trip_ids}
        
        # Process results as they complete
        for future in concurrent.futures.as_completed(future_to_trip):
            trip_id = future_to_trip[future]
            try:
                trip_stats = future.result()
                # Aggregate statistics
                stats["updated"] += trip_stats["updated"]
                stats["skipped"] += trip_stats["skipped"]
                stats["errors"] += trip_stats["errors"]
                stats["created"] += trip_stats["created"]
                
                for field, count in trip_stats["updated_fields"].items():
                    stats["updated_fields"][field] += count
                    
                for reason, count in trip_stats["reasons"].items():
                    stats["reasons"][reason] += count
                    
            except Exception as e:
                stats["errors"] += 1
                print(f"Exception processing trip {trip_id}: {e}")
    
    session_local.close()
    
    # Prepare detailed feedback message
    if stats["updated"] > 0:
        message = f"Updated {stats['updated']} trips ({stats['created']} new, {stats['skipped']} skipped, {stats['errors']} errors)"
        
        # Add detailed field statistics if any fields were updated
        if stats["updated_fields"]:
            message += "<br><br>Fields updated:<ul>"
            for field, count in stats["updated_fields"].most_common():
                message += f"<li>{field}: {count} trips</li>"
            message += "</ul>"
            
        # Add detailed reason statistics
        if stats["reasons"]:
            message += "<br>Reasons for updates:<ul>"
            for reason, count in stats["reasons"].most_common():
                message += f"<li>{reason}: {count} trips</li>"
            message += "</ul>"
            
        return message
    else:
        return "No trips were updated. All trips are up to date."

@app.route("/export_trips")
def export_trips():
    """
    Export filtered trips to XLSX, merging with DB data (including trip_time, completed_by,
    coordinate_count (log count), status, route_quality, expected_trip_quality, and lack_of_accuracy).
    Supports operator-based filtering and range filtering for trip_time, log_count, and also for:
      - Short Segments (<1km)
      - Medium Segments (1-5km)
      - Long Segments (>5km)
      - Short Dist Total
      - Medium Dist Total
      - Long Dist Total
      - Max Segment Dist
      - Avg Segment Dist
    """
    session_local = db_session()
    # Basic filters from the request
    filters = {
        "driver": request.args.get("driver"),
        "trip_id": request.args.get("trip_id"),
        "model": request.args.get("model"),
        "ram": request.args.get("ram"),
        "carrier": request.args.get("carrier"),
        "variance_min": request.args.get("variance_min"),
        "variance_max": request.args.get("variance_max"),
        "export_name": request.args.get("export_name", "exported_trips"),
        "route_quality": request.args.get("route_quality", "").strip(),
        "trip_issues": request.args.get("trip_issues", "").strip(),
        "lack_of_accuracy": request.args.get("lack_of_accuracy", "").strip(),
        "tags": request.args.get("tags", "").strip()
    }
    # Filters with operator strings for trip_time and log_count
    trip_time = request.args.get("trip_time", "").strip()
    trip_time_op = request.args.get("trip_time_op", "equal").strip()
    completed_by_filter = request.args.get("completed_by", "").strip()
    log_count = request.args.get("log_count", "").strip()
    log_count_op = request.args.get("log_count_op", "equal").strip()
    status_filter = request.args.get("status", "").strip()
    
    # New range filter parameters for trip_time and log_count
    trip_time_min = request.args.get("trip_time_min", "").strip()
    trip_time_max = request.args.get("trip_time_max", "").strip()
    log_count_min = request.args.get("log_count_min", "").strip()
    log_count_max = request.args.get("log_count_max", "").strip()

    # NEW: New query parameters for segment analysis fields
    medium_segments = request.args.get("medium_segments", "").strip()
    medium_segments_op = request.args.get("medium_segments_op", "equal").strip()
    long_segments = request.args.get("long_segments", "").strip()
    long_segments_op = request.args.get("long_segments_op", "equal").strip()
    short_dist_total = request.args.get("short_dist_total", "").strip()
    short_dist_total_op = request.args.get("short_dist_total_op", "equal").strip()
    medium_dist_total = request.args.get("medium_dist_total", "").strip()
    medium_dist_total_op = request.args.get("medium_dist_total_op", "equal").strip()
    long_dist_total = request.args.get("long_dist_total", "").strip()
    long_dist_total_op = request.args.get("long_dist_total_op", "equal").strip()
    max_segment_distance = request.args.get("max_segment_distance", "").strip()
    max_segment_distance_op = request.args.get("max_segment_distance_op", "equal").strip()
    avg_segment_distance = request.args.get("avg_segment_distance", "").strip()
    avg_segment_distance_op = request.args.get("avg_segment_distance_op", "equal").strip()

    # NEW: Expected Trip Quality filter
    expected_trip_quality_filter = request.args.get("expected_trip_quality", "").strip()

    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)
    merged = []

    # Date range filtering code
    start_date_param = request.args.get('start_date')
    end_date_param = request.args.get('end_date')
    if start_date_param and end_date_param:
        start_date_filter = None
        end_date_filter = None
        for fmt in ["%Y-%m-%d", "%d-%m-%Y"]:
            try:
                start_date_filter = datetime.strptime(start_date_param, fmt)
                end_date_filter = datetime.strptime(end_date_param, fmt)
                break
            except ValueError:
                continue
        if start_date_filter and end_date_filter:
            filtered_data = []
            for row in excel_data:
                if row.get('time'):
                    try:
                        row_time = row['time']
                        if isinstance(row_time, str):
                            row_time = datetime.strptime(row_time, "%Y-%m-%d %H:%M:%S")
                        if start_date_filter.date() <= row_time.date() < end_date_filter.date():
                            filtered_data.append(row)
                    except Exception:
                        continue
            excel_data = filtered_data

    all_times = []
    for row in excel_data:
        if row.get('time'):
            try:
                row_time = row['time']
                if isinstance(row_time, str):
                    row_time = datetime.strptime(row_time, "%Y-%m-%d %H:%M:%S")
                all_times.append(row_time)
            except Exception:
                continue
    min_date = min(all_times) if all_times else None
    max_date = max(all_times) if all_times else None

    # Basic Excel filters
    if filters["driver"]:
        excel_data = [row for row in excel_data if str(row.get("UserName", "")).strip() == filters["driver"]]
    if filters["trip_id"]:
        try:
            tid = int(filters["trip_id"])
            excel_data = [row for row in excel_data if row.get("tripId") == tid]
        except ValueError:
            pass
    if filters["model"]:
        excel_data = [row for row in excel_data if str(row.get("model", "")).strip() == filters["model"]]
    if filters["ram"]:
        excel_data = [row for row in excel_data if str(row.get("RAM", "")).strip() == filters["ram"]]
    if filters["carrier"]:
        excel_data = [row for row in excel_data if str(row.get("carrier", "")).strip().lower() == filters["carrier"].lower()]

    # Merge Excel data with DB records
    excel_trip_ids = [row.get("tripId") for row in excel_data if row.get("tripId")]
    if filters["tags"]:
        query = db_session.query(Trip).filter(Trip.trip_id.in_(excel_trip_ids)).join(Trip.tags).filter(Tag.name.ilike('%' + filters["tags"] + '%'))
        db_trips = query.all()
        filtered_trip_ids = [trip.trip_id for trip in db_trips]
        excel_data = [r for r in excel_data if r.get("tripId") in filtered_trip_ids]
        db_trip_map = {trip.trip_id: trip for trip in db_trips}
    else:
        trip_issues_filter = filters.get("trip_issues", "")
        query = db_session.query(Trip).filter(Trip.trip_id.in_(excel_trip_ids))
        if trip_issues_filter:
            query = query.join(Trip.tags).filter(Tag.name.ilike('%' + trip_issues_filter + '%'))
        db_trips = query.all()
        db_trip_map = {trip.trip_id: trip for trip in db_trips}

    for row in excel_data:
        trip_id = row.get("tripId")
        db_trip = db_trip_map.get(trip_id)
        if db_trip:
            try:
                md = float(db_trip.manual_distance)
            except (TypeError, ValueError):
                md = None
            try:
                cd = float(db_trip.calculated_distance)
            except (TypeError, ValueError):
                cd = None
            row["route_quality"] = db_trip.route_quality or ""
            row["manual_distance"] = md if md is not None else ""
            row["calculated_distance"] = cd if cd is not None else ""
            if md and cd and md != 0:
                pct = (cd / md) * 100
                row["distance_percentage"] = f"{pct:.2f}%"
                variance = abs(cd - md) / md * 100
                row["variance"] = variance
            else:
                row["distance_percentage"] = "N/A"
                row["variance"] = None
            # Other fields

            row["trip_time"] = db_trip.trip_time if db_trip.trip_time is not None else ""
            row["completed_by"] = db_trip.completed_by if db_trip.completed_by is not None else ""
            row["coordinate_count"] = db_trip.coordinate_count if db_trip.coordinate_count is not None else ""
            row["status"] = db_trip.status if db_trip.status is not None else ""
            row["lack_of_accuracy"] = db_trip.lack_of_accuracy if db_trip.lack_of_accuracy is not None else ""
            row["trip_issues"] = ", ".join([tag.name for tag in db_trip.tags]) if db_trip.tags else ""
            row["tags"] = row["trip_issues"]
            row["expected_trip_quality"] = str(db_trip.expected_trip_quality) if db_trip.expected_trip_quality is not None else "N/A"
            # Include the segment analysis fields
            row["medium_segments_count"] = db_trip.medium_segments_count
            row["long_segments_count"] = db_trip.long_segments_count
            row["short_segments_distance"] = db_trip.short_segments_distance
            row["medium_segments_distance"] = db_trip.medium_segments_distance
            row["long_segments_distance"] = db_trip.long_segments_distance
            row["max_segment_distance"] = db_trip.max_segment_distance
            row["avg_segment_distance"] = db_trip.avg_segment_distance

        else:
            row["route_quality"] = ""
            row["manual_distance"] = ""
            row["calculated_distance"] = ""
            row["distance_percentage"] = "N/A"
            row["variance"] = None
            row["trip_time"] = ""
            row["completed_by"] = ""
            row["coordinate_count"] = ""
            row["status"] = ""
            row["lack_of_accuracy"] = ""
            row["trip_issues"] = ""
            row["tags"] = ""
            row["expected_trip_quality"] = "N/A"
            row["medium_segments_count"] = None
            row["long_segments_count"] = None
            row["short_segments_distance"] = None
            row["medium_segments_distance"] = None
            row["long_segments_distance"] = None
            row["max_segment_distance"] = None
            row["avg_segment_distance"] = None

        merged.append(row)

    # Additional variance filters
    if filters["variance_min"]:
        try:
            vmin = float(filters["variance_min"])
            merged = [r for r in merged if r.get("variance") is not None and r["variance"] >= vmin]
        except ValueError:
            pass
    if filters["variance_max"]:
        try:
            vmax = float(filters["variance_max"])
            merged = [r for r in merged if r.get("variance") is not None and r["variance"] <= vmax]
        except ValueError:
            pass

    # Now filter by route_quality based on merged (DB) value.
    if filters["route_quality"]:
        rq_filter = filters["route_quality"].lower().strip()
        if rq_filter == "not assigned":
            merged = [r for r in merged if str(r.get("route_quality", "")).strip() == ""]
        else:
            merged = [r for r in merged if str(r.get("route_quality", "")).strip().lower() == rq_filter]
    
    # Apply lack_of_accuracy filter after merging
    if filters["lack_of_accuracy"]:
        lo_filter = filters["lack_of_accuracy"].lower()
        if lo_filter in ['true', 'yes', '1']:
            merged = [r for r in merged if r.get("lack_of_accuracy") is True]
        elif lo_filter in ['false', 'no', '0']:
            merged = [r for r in merged if r.get("lack_of_accuracy") is False]

    # Apply expected_trip_quality filter if provided
    if expected_trip_quality_filter:
        merged = [r for r in merged if str(r.get("expected_trip_quality", "")).strip().lower() == expected_trip_quality_filter.lower()]

    # Helper functions for numeric comparisons
    def normalize_op(op):
        op = op.lower().strip()
        mapping = {
            "equal": "=",
            "equals": "=",
            "=": "=",
            "less than": "<",
            "more than": ">",
            "less than or equal": "<=",
            "less than or equal to": "<=",
            "more than or equal": ">=",
            "more than or equal to": ">="
        }
        return mapping.get(op, "=")

    def compare(value, op, threshold):
        op = normalize_op(op)
        if op == "=":
            return value == threshold
        elif op == "<":
            return value < threshold
        elif op == ">":
            return value > threshold
        elif op == "<=":
            return value <= threshold
        elif op == ">=":
            return value >= threshold
        return False

    # Filter by trip_time
    if trip_time_min or trip_time_max:
        if trip_time_min:
            try:
                tt_min = float(trip_time_min)
                merged = [r for r in merged if r.get("trip_time") not in (None, "") and float(r.get("trip_time")) >= tt_min]
            except ValueError:
                pass
        if trip_time_max:
            try:
                tt_max = float(trip_time_max)
                merged = [r for r in merged if r.get("trip_time") not in (None, "") and float(r.get("trip_time")) <= tt_max]
            except ValueError:
                pass
    elif trip_time:
        try:
            tt_value = float(trip_time)
            merged = [r for r in merged if r.get("trip_time") not in (None, "") and compare(float(r.get("trip_time")), trip_time_op, tt_value)]
        except ValueError:
            pass

    # Filter by completed_by (case-insensitive)
    if completed_by_filter:
        merged = [r for r in merged if r.get("completed_by") and str(r.get("completed_by")).strip().lower() == completed_by_filter.lower()]

    # Filter by log_count
    if log_count_min or log_count_max:
        if log_count_min:
            try:
                lc_min = int(log_count_min)
                merged = [r for r in merged if r.get("coordinate_count") not in (None, "") and int(r.get("coordinate_count")) >= lc_min]
            except ValueError:
                pass
        if log_count_max:
            try:
                lc_max = int(log_count_max)
                merged = [r for r in merged if r.get("coordinate_count") not in (None, "") and int(r.get("coordinate_count")) <= lc_max]
            except ValueError:
                pass
    elif log_count:
        try:
            lc_value = int(log_count)
            merged = [r for r in merged if r.get("coordinate_count") not in (None, "") and compare(int(r.get("coordinate_count")), log_count_op, lc_value)]
        except ValueError:
            pass

    # Operator filtering for segment analysis fields:

    # Medium Segments Count
    if medium_segments:
        try:
            ms_value = int(medium_segments)
            merged = [r for r in merged if compare(int(r.get("medium_segments_count") or 0), medium_segments_op, ms_value)]
        except ValueError:
            pass

    # Long Segments Count
    if long_segments:
        try:
            ls_value = int(long_segments)
            merged = [r for r in merged if compare(int(r.get("long_segments_count") or 0), long_segments_op, ls_value)]
        except ValueError:
            pass

    # Short Distance Total
    if short_dist_total:
        try:
            sdt_value = float(short_dist_total)
            merged = [r for r in merged if compare(float(r.get("short_segments_distance") or 0.0), short_dist_total_op, sdt_value)]
        except ValueError:
            pass

    # Medium Distance Total
    if medium_dist_total:
        try:
            mdt_value = float(medium_dist_total)
            merged = [r for r in merged if compare(float(r.get("medium_segments_distance") or 0.0), medium_dist_total_op, mdt_value)]
        except ValueError:
            pass

    # Long Distance Total
    if long_dist_total:
        try:
            ldt_value = float(long_dist_total)
            merged = [r for r in merged if compare(float(r.get("long_segments_distance") or 0.0), long_dist_total_op, ldt_value)]
        except ValueError:
            pass

    # Max Segment Distance
    if max_segment_distance:
        try:
            msd_value = float(max_segment_distance)
            merged = [r for r in merged if compare(float(r.get("max_segment_distance") or 0.0), max_segment_distance_op, msd_value)]
        except ValueError:
            pass

    # Avg Segment Distance
    if avg_segment_distance:
        try:
            asd_value = float(avg_segment_distance)
            merged = [r for r in merged if compare(float(r.get("avg_segment_distance") or 0.0), avg_segment_distance_op, asd_value)]
        except ValueError:
            pass

    # Filter by status
    if status_filter:
        status_lower = status_filter.lower().strip()
        if status_lower in ("empty", "not assigned"):
            merged = [r for r in merged if not r.get("status") or str(r.get("status")).strip() == ""]
        else:
            merged = [r for r in merged if r.get("status") and str(r.get("status")).strip().lower() == status_lower]

    wb = Workbook()
    ws = wb.active
    if merged:
        headers = list(merged[0].keys())
        ws.append(headers)
        for row in merged:
            ws.append([row.get(col) for col in headers])
    else:
        ws.append(["No data found"])

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"{filters['export_name']}.xlsx"
    session_local.close()
    return send_file(
        file_stream,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )










# ---------------------------
# Dashboard (Analytics) - Consolidated by User, with Date Range
# ---------------------------
@app.route("/")
def analytics():
    """
    Main dashboard page with a toggle for:
      - data_scope = 'all'   => analyze ALL trips in DB
      - data_scope = 'excel' => only the trip IDs in the current data.xlsx
    We store the user's choice in the session so it persists until changed.
    """
    session_local = db_session()

    # 1) Check if user provided data_scope in request
    if "data_scope" in request.args:
        chosen_scope = request.args.get("data_scope", "all")
        flask_session["data_scope"] = chosen_scope
    else:
        chosen_scope = flask_session.get("data_scope", "all")  # default 'all'

    # 2) Additional filters for analytics page
    driver_filter = request.args.get("driver", "").strip()
    carrier_filter = request.args.get("carrier", "").strip()

    # 3) Load Excel data & merge route_quality from DB
    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)
    excel_trip_ids = [r["tripId"] for r in excel_data if r.get("tripId")]
    session_local = db_session()
    db_trips_for_excel = session_local.query(Trip).filter(Trip.trip_id.in_(excel_trip_ids)).all()
    db_map = {t.trip_id: t for t in db_trips_for_excel}
    for row in excel_data:
        trip_id = row.get("tripId")
        if trip_id in db_map:
            row["route_quality"] = db_map[trip_id].route_quality or ""
        else:
            row.setdefault("route_quality", "")

    # 4) Decide which DB trips to analyze for distance accuracy
    if chosen_scope == "excel":
        trips_db = db_trips_for_excel
    else:
        trips_db = session_local.query(Trip).all()

    # 5) Compute distance accuracy
    correct = 0
    incorrect = 0
    for trip in trips_db:
        try:
            md = float(trip.manual_distance)
            cd = float(trip.calculated_distance)
            if md and md != 0:
                if abs(cd - md) / md <= 0.2:
                    correct += 1
                else:
                    incorrect += 1
        except:
            pass
    total_trips = correct + incorrect
    if total_trips > 0:
        correct_pct = correct / total_trips * 100
        incorrect_pct = incorrect / total_trips * 100
    else:
        correct_pct = 0
        incorrect_pct = 0

    # 6) Build a filtered "excel-like" dataset for the user-level charts
    if chosen_scope == "excel":
        # Just the real Excel data
        filtered_excel_data = excel_data[:]
    else:
        # All DB trips, but we create placeholders if a trip isn't in Excel
        all_db = trips_db
        excel_map = {r["tripId"]: r for r in excel_data if r.get("tripId")}
        all_data_rows = []
        for tdb in all_db:
            if tdb.trip_id in excel_map:
                row_copy = dict(excel_map[tdb.trip_id])
                row_copy["route_quality"] = tdb.route_quality or ""
            else:
                row_copy = {
                    "tripId": tdb.trip_id,
                    "UserName": "",
                    "carrier": "",
                    "Android Version": "",
                    "manufacturer": "",
                    "model": "",
                    "RAM": "",
                    "route_quality": tdb.route_quality or ""
                }
            all_data_rows.append(row_copy)
        filtered_excel_data = all_data_rows

    # 7) Apply driver & carrier filters
    if driver_filter:
        filtered_excel_data = [r for r in filtered_excel_data if str(r.get("UserName","")).strip() == driver_filter]

    if carrier_filter:
        # user picked one of the 4 carriers => keep only matching normalized
        new_list = []
        for row in filtered_excel_data:
            norm_car = normalize_carrier(row.get("carrier",""))
            if norm_car == carrier_filter:
                new_list.append(row)
        filtered_excel_data = new_list

    # 8) Consolidate user-latest for charts
    user_latest = {}
    for row in filtered_excel_data:
        user = str(row.get("UserName","")).strip()
        if user:
            user_latest[user] = row
    consolidated_rows = list(user_latest.values())

    # Prepare chart data
    carrier_counts = {}
    os_counts = {}
    manufacturer_counts = {}
    model_counts = {}

    for row in consolidated_rows:
        c = normalize_carrier(row.get("carrier",""))
        carrier_counts[c] = carrier_counts.get(c,0)+1

        osv = row.get("Android Version")
        osv = str(osv) if osv is not None else "Unknown"
        os_counts[osv] = os_counts.get(osv, 0) + 1

        manu = row.get("manufacturer","Unknown")
        manufacturer_counts[manu] = manufacturer_counts.get(manu,0)+1

        mdl = row.get("model","UnknownModel")
        model_counts[mdl] = model_counts.get(mdl,0)+1

    total_users = len(consolidated_rows)
    device_usage = []
    for mdl, cnt in model_counts.items():
        pct = (cnt / total_users * 100) if total_users else 0
        device_usage.append({"model": mdl, "count": cnt, "percentage": round(pct,2)})

    # Build user_data for High/Low/Other
    user_data = {}
    for row in filtered_excel_data:
        user = str(row.get("UserName","")).strip()
        if not user:
            continue
        if user not in user_data:
            user_data[user] = {
                "total_trips": 0,
                "No Logs Trips": 0,
                "Trip Points Only Exist": 0,
                "Low": 0,
                "Moderate": 0,
                "High": 0,
                "Other": 0
            }
        user_data[user]["total_trips"] += 1
        q = row.get("route_quality", "")
        if q in ["No Logs Trips", "Trip Points Only Exist", "Low", "Moderate", "High"]:
            user_data[user][q] += 1
        else:
            user_data[user]["Other"] += 1

    # Quality analysis
    high_quality_models = {}
    low_quality_models = {}
    high_quality_android = {}
    low_quality_android = {}
    high_quality_ram = {}
    low_quality_ram = {}

    sensor_cols = [
        "Fingerprint Sensor","Accelerometer","Gyro",
        "Proximity Sensor","Compass","Barometer",
        "Background Task Killing Tendency"
    ]
    high_quality_sensors = {s:0 for s in sensor_cols}
    total_high_quality = 0

    for row in filtered_excel_data:
        q = row.get("route_quality","")
        mdl = row.get("model","UnknownModel")
        av = row.get("Android Version","Unknown")
        ram = row.get("RAM","")
        if q == "High":
            total_high_quality +=1
            high_quality_models[mdl] = high_quality_models.get(mdl,0)+1
            high_quality_android[av] = high_quality_android.get(av,0)+1
            high_quality_ram[ram] = high_quality_ram.get(ram,0)+1
            for sensor in sensor_cols:
                val = row.get(sensor,"")
                if (isinstance(val,str) and val.lower()=="true") or (val is True):
                    high_quality_sensors[sensor]+=1
        elif q == "Low":
            low_quality_models[mdl] = low_quality_models.get(mdl,0)+1
            low_quality_android[av] = low_quality_android.get(av,0)+1
            low_quality_ram[ram] = low_quality_ram.get(ram,0)+1

    session_local.close()

    # Build driver list for the dropdown
    all_drivers = sorted({str(r.get("UserName","")).strip() for r in excel_data if r.get("UserName")})
    carriers_for_dropdown = ["Vodafone","Orange","Etisalat","We"]

    return render_template(
        "analytics.html",
        data_scope=chosen_scope,
        driver_filter=driver_filter,
        carrier_filter=carrier_filter,
        drivers=all_drivers,
        carriers_for_dropdown=carriers_for_dropdown,
        carrier_counts=carrier_counts,
        os_counts=os_counts,
        manufacturer_counts=manufacturer_counts,
        device_usage=device_usage,
        total_trips=total_trips,
        correct_pct=correct_pct,
        incorrect_pct=incorrect_pct,
        user_data=user_data,
        high_quality_models=high_quality_models,
        low_quality_models=low_quality_models,
        high_quality_android=high_quality_android,
        low_quality_android=low_quality_android,
        high_quality_ram=high_quality_ram,
        low_quality_ram=low_quality_ram,
        high_quality_sensors=high_quality_sensors,
        total_high_quality=total_high_quality
    )


# ---------------------------
# Trips Page with Variance, Pagination, etc.
# ---------------------------
@app.route("/trips")
def trips():
    """
    Trips page with filtering (including trip_time, completed_by, log_count, status, route_quality,
    lack_of_accuracy, expected_trip_quality, segment analysis filters, and tags) with operator support
    for trip_time and log_count) and pagination.
    """
    session_local = db_session()
    page = request.args.get("page", type=int, default=1)
    page_size = 100
    if page < 1:
        page = 1

    # Extract only non-empty filter parameters
    filters = {}
    for key, value in request.args.items():
        if value and value.strip():
            filters[key] = value.strip()

    # Extract basic filter parameters
    driver_filter = filters.get("driver", "")
    trip_id_search = filters.get("trip_id", "")
    route_quality_filter = filters.get("route_quality", "")
    model_filter = filters.get("model", "")
    ram_filter = filters.get("ram", "")
    carrier_filter = filters.get("carrier", "")
    variance_min = float(filters["variance_min"]) if "variance_min" in filters else None
    variance_max = float(filters["variance_max"]) if "variance_max" in filters else None
    trip_time_filter = filters.get("trip_time", "")
    trip_time_op = filters.get("trip_time_op", "equal")
    completed_by_filter = filters.get("completed_by", "")
    log_count_filter = filters.get("log_count", "")
    log_count_op = filters.get("log_count_op", "equal")
    status_filter = filters.get("status", "completed")
    lack_of_accuracy_filter = filters.get("lack_of_accuracy", "").lower()
    tags_filter = filters.get("tags", "")

    # Expected trip quality filter
    expected_trip_quality_filter = filters.get("expected_trip_quality", "")

    # Extract range filters for trip_time and log_count
    trip_time_min = filters.get("trip_time_min", "")
    trip_time_max = filters.get("trip_time_max", "")
    log_count_min = filters.get("log_count_min", "")
    log_count_max = filters.get("log_count_max", "")

    # Extract segment analysis filter parameters
    medium_segments = filters.get("medium_segments", "")
    medium_segments_op = filters.get("medium_segments_op", "equal")
    long_segments = filters.get("long_segments", "")
    long_segments_op = filters.get("long_segments_op", "equal")
    short_dist_total = filters.get("short_dist_total", "")
    short_dist_total_op = filters.get("short_dist_total_op", "equal")
    medium_dist_total = filters.get("medium_dist_total", "")
    medium_dist_total_op = filters.get("medium_dist_total_op", "equal")
    long_dist_total = filters.get("long_dist_total", "")
    long_dist_total_op = filters.get("long_dist_total_op", "equal")
    max_segment_distance = filters.get("max_segment_distance", "")
    max_segment_distance_op = filters.get("max_segment_distance_op", "equal")
    avg_segment_distance = filters.get("avg_segment_distance", "")
    avg_segment_distance_op = filters.get("avg_segment_distance_op", "equal")

    # Define helper functions for numeric comparisons
    def normalize_op(op):
        op = op.lower().strip()
        mapping = {
            "equal": "=",
            "equals": "=",
            "=": "=",
            "less than": "<",
            "more than": ">",
            "less than or equal": "<=",
            "less than or equal to": "<=",
            "more than or equal": ">=",
            "more than or equal to": ">="
        }
        return mapping.get(op, "=")

    def compare(value, op, threshold):
        op = normalize_op(op)
        if op == "=":
            return value == threshold
        elif op == "<":
            return value < threshold
        elif op == ">":
            return value > threshold
        elif op == "<=":
            return value <= threshold
        elif op == ">=":
            return value >= threshold
        return False

    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)
    merged = []

    # Date range filtering code (omitted here for brevity)
    start_date_param = request.args.get('start_date')
    end_date_param = request.args.get('end_date')
    if start_date_param and end_date_param:
        start_date_filter = None
        end_date_filter = None
        for fmt in ["%Y-%m-%d", "%d-%m-%Y"]:
            try:
                start_date_filter = datetime.strptime(start_date_param, fmt)
                end_date_filter = datetime.strptime(end_date_param, fmt)
                break
            except ValueError:
                continue
        if start_date_filter and end_date_filter:
            filtered_data = []
            for row in excel_data:
                if row.get('time'):
                    try:
                        row_time = row['time']
                        if isinstance(row_time, str):
                            row_time = datetime.strptime(row_time, "%Y-%m-%d %H:%M:%S")
                        if start_date_filter.date() <= row_time.date() < end_date_filter.date():
                            filtered_data.append(row)
                    except Exception:
                        continue
            excel_data = filtered_data

    all_times = []
    for row in excel_data:
        if row.get('time'):
            try:
                row_time = row['time']
                if isinstance(row_time, str):
                    row_time = datetime.strptime(row_time, "%Y-%m-%d %H:%M:%S")
                all_times.append(row_time)
            except Exception:
                continue
    min_date = min(all_times) if all_times else None
    max_date = max(all_times) if all_times else None

    if driver_filter:
        excel_data = [r for r in excel_data if str(r.get("UserName", "")).strip() == driver_filter]
    if trip_id_search:
        try:
            tid = int(trip_id_search)
            excel_data = [r for r in excel_data if r.get("tripId") == tid]
        except ValueError:
            pass
    if model_filter:
        excel_data = [r for r in excel_data if str(r.get("model", "")).strip() == model_filter]
    if ram_filter:
        excel_data = [r for r in excel_data if str(r.get("RAM", "")).strip() == ram_filter]
    if carrier_filter:
        new_list = []
        for row in excel_data:
            norm_car = normalize_carrier(row.get("carrier", ""))
            if norm_car == carrier_filter:
                new_list.append(row)
        excel_data = new_list

    excel_trip_ids = [r["tripId"] for r in excel_data if r.get("tripId")]
    if tags_filter:
        db_trips = session_local.query(Trip).filter(Trip.trip_id.in_(excel_trip_ids)).join(Trip.tags).filter(Tag.name.ilike('%' + tags_filter + '%')).all()
        filtered_trip_ids = [trip.trip_id for trip in db_trips]
        excel_data = [r for r in excel_data if r.get("tripId") in filtered_trip_ids]
    else:
        db_trips = session_local.query(Trip).filter(Trip.trip_id.in_(excel_trip_ids)).all()
    
    db_map = {t.trip_id: t for t in db_trips}
    for row in excel_data:
        tdb = db_map.get(row["tripId"])
        if tdb:
            try:
                md = float(tdb.manual_distance)
            except:
                md = None
            try:
                cd = float(tdb.calculated_distance)
            except:
                cd = None
            row["route_quality"] = tdb.route_quality or ""
            row["manual_distance"] = md if md is not None else ""
            row["calculated_distance"] = cd if cd is not None else ""
            row["trip_time"] = tdb.trip_time if tdb.trip_time is not None else ""
            row["completed_by"] = tdb.completed_by if tdb.completed_by is not None else ""
            row["coordinate_count"] = tdb.coordinate_count if tdb.coordinate_count is not None else ""
            row["status"] = tdb.status if tdb.status is not None else ""
            row["lack_of_accuracy"] = tdb.lack_of_accuracy if tdb.lack_of_accuracy is not None else ""
            row["short_segments_count"] = tdb.short_segments_count
            row["medium_segments_count"] = tdb.medium_segments_count
            row["long_segments_count"] = tdb.long_segments_count
            row["short_segments_distance"] = tdb.short_segments_distance
            row["medium_segments_distance"] = tdb.medium_segments_distance
            row["long_segments_distance"] = tdb.long_segments_distance
            row["max_segment_distance"] = tdb.max_segment_distance
            row["avg_segment_distance"] = tdb.avg_segment_distance
            row["trip_issues"] = ", ".join([tag.name for tag in tdb.tags]) if tdb.tags else ""
            row["tags"] = row["trip_issues"]
            if md and cd and md != 0:
                pct = (cd / md) * 100
                row["distance_percentage"] = f"{pct:.2f}%"
                var = abs(cd - md) / md * 100
                row["variance"] = var
            else:
                row["distance_percentage"] = "N/A"
                row["variance"] = None
            row["expected_trip_quality"] = tdb.expected_trip_quality if tdb.expected_trip_quality is not None else "N/A"
        else:
            row["route_quality"] = ""
            row["manual_distance"] = ""
            row["calculated_distance"] = ""
            row["trip_time"] = ""
            row["completed_by"] = ""
            row["coordinate_count"] = ""
            row["status"] = ""
            row["lack_of_accuracy"] = ""
            row["short_segments_count"] = None
            row["medium_segments_count"] = None
            row["long_segments_count"] = None
            row["short_segments_distance"] = None
            row["medium_segments_distance"] = None
            row["long_segments_distance"] = None
            row["max_segment_distance"] = None
            row["avg_segment_distance"] = None
            row["trip_issues"] = ""
            row["tags"] = ""
            # Set expected_trip_quality to N/A if no DB record exists
            row["expected_trip_quality"] = "N/A"
        merged.append(row)

    # Apply route_quality filter after merging
    if route_quality_filter:
        rq_filter = route_quality_filter.lower().strip()
        if rq_filter == "not assigned":
            excel_data = [r for r in excel_data if str(r.get("route_quality", "")).strip() == ""]
        else:
            excel_data = [r for r in excel_data if str(r.get("route_quality", "")).strip().lower() == rq_filter]
    
    # Apply lack_of_accuracy filter after merging
    if lack_of_accuracy_filter:
        if lack_of_accuracy_filter in ['true', 'yes', '1']:
            excel_data = [r for r in excel_data if r.get("lack_of_accuracy") is True]
        elif lack_of_accuracy_filter in ['false', 'no', '0']:
            excel_data = [r for r in excel_data if r.get("lack_of_accuracy") is False]
    
    if variance_min is not None:
        excel_data = [r for r in excel_data if r.get("variance") is not None and r["variance"] >= variance_min]
    if variance_max is not None:
        excel_data = [r for r in excel_data if r.get("variance") is not None and r["variance"] <= variance_max]
    
    # Apply expected_trip_quality filter if provided
    if expected_trip_quality_filter:
        excel_data = [r for r in excel_data if str(r.get("expected_trip_quality", "")).strip().lower() == expected_trip_quality_filter.lower()]

    # --- Apply segment analysis filters ---
    if medium_segments:
        try:
            ms_value = int(medium_segments)
            excel_data = [r for r in excel_data if compare(int(r.get("medium_segments_count") or 0), medium_segments_op, ms_value)]
        except ValueError:
            pass

    if long_segments:
        try:
            ls_value = int(long_segments)
            excel_data = [r for r in excel_data if compare(int(r.get("long_segments_count") or 0), long_segments_op, ls_value)]
        except ValueError:
            pass

    if short_dist_total:
        try:
            sdt_value = float(short_dist_total)
            excel_data = [r for r in excel_data if compare(float(r.get("short_segments_distance") or 0.0), short_dist_total_op, sdt_value)]
        except ValueError:
            pass

    if medium_dist_total:
        try:
            mdt_value = float(medium_dist_total)
            excel_data = [r for r in excel_data if compare(float(r.get("medium_segments_distance") or 0.0), medium_dist_total_op, mdt_value)]
        except ValueError:
            pass

    if long_dist_total:
        try:
            ldt_value = float(long_dist_total)
            excel_data = [r for r in excel_data if compare(float(r.get("long_segments_distance") or 0.0), long_dist_total_op, ldt_value)]
        except ValueError:
            pass

    if max_segment_distance:
        try:
            msd_value = float(max_segment_distance)
            excel_data = [r for r in excel_data if compare(float(r.get("max_segment_distance") or 0.0), max_segment_distance_op, msd_value)]
        except ValueError:
            pass

    if avg_segment_distance:
        try:
            asd_value = float(avg_segment_distance)
            excel_data = [r for r in excel_data if compare(float(r.get("avg_segment_distance") or 0.0), avg_segment_distance_op, asd_value)]
        except ValueError:
            pass

    # --- Apply trip_time filters ---
    if trip_time_min or trip_time_max:
        if trip_time_min:
            try:
                tt_min = float(trip_time_min)
                excel_data = [r for r in excel_data if r.get("trip_time") not in (None, "") and float(r.get("trip_time")) >= tt_min]
            except ValueError:
                pass
        if trip_time_max:
            try:
                tt_max = float(trip_time_max)
                excel_data = [r for r in excel_data if r.get("trip_time") not in (None, "") and float(r.get("trip_time")) <= tt_max]
            except ValueError:
                pass
    elif trip_time_filter:
        try:
            tt_value = float(trip_time_filter)
            excel_data = [r for r in excel_data if r.get("trip_time") not in (None, "") and compare(float(r.get("trip_time")), trip_time_op, tt_value)]
        except ValueError:
            pass

    if completed_by_filter:
        excel_data = [r for r in excel_data if r.get("completed_by") and str(r.get("completed_by")).strip().lower() == completed_by_filter.lower()]

    if log_count_min or log_count_max:
        if log_count_min:
            try:
                lc_min = int(log_count_min)
                excel_data = [r for r in excel_data if r.get("coordinate_count") not in (None, "") and int(r.get("coordinate_count")) >= lc_min]
            except ValueError:
                pass
        if log_count_max:
            try:
                lc_max = int(log_count_max)
                excel_data = [r for r in excel_data if r.get("coordinate_count") not in (None, "") and int(r.get("coordinate_count")) <= lc_max]
            except ValueError:
                pass
    elif log_count_filter:
        try:
            lc_value = int(log_count_filter)
            excel_data = [r for r in excel_data if r.get("coordinate_count") not in (None, "") and compare(int(r.get("coordinate_count")), log_count_op, lc_value)]
        except ValueError:
            pass

    if status_filter:
        status_lower = status_filter.lower().strip()
        if status_lower in ("empty", "not assigned"):
            excel_data = [r for r in excel_data if not r.get("status") or str(r.get("status")).strip() == ""]
        else:
            excel_data = [r for r in excel_data if r.get("status") and str(r.get("status")).strip().lower() == status_lower]

    total_rows = len(excel_data)
    total_pages = (total_rows + page_size - 1) // page_size if total_rows else 1
    if page > total_pages and total_pages > 0:
        page = total_pages
    start = (page - 1) * page_size
    end = start + page_size
    page_data = excel_data[start:end]

    all_tags = session_local.query(Tag).all()
    tags_for_dropdown = [tag.name for tag in all_tags]

    session_local.close()

    all_excel = load_excel_data(excel_path)
    statuses = sorted(set(r.get("status", "").strip() for r in all_excel if r.get("status") and r.get("status").strip()))
    completed_by_options = sorted(set(r.get("completed_by", "").strip() for r in all_excel if r.get("completed_by") and r.get("completed_by").strip()))
    model_set = {}
    for r in all_excel:
        m = r.get("model", "").strip()
        device = r.get("Device Name", "").strip() if r.get("Device Name") else ""
        if m:
            display = m
            if device:
                display += " - " + device
            model_set[m] = display
    models_options = sorted(model_set.items(), key=lambda x: x[1])

    if not statuses:
        session_temp = db_session()
        statuses = sorted(set(row[0].strip() for row in session_temp.query(Trip.status).filter(Trip.status != None).distinct().all() if row[0] and row[0].strip()))
        session_temp.close()
    if not completed_by_options:
        session_temp = db_session()
        completed_by_options = sorted(set(row[0].strip() for row in session_temp.query(Trip.completed_by).filter(Trip.completed_by != None).distinct().all() if row[0] and row[0].strip()))
        session_temp.close()
    drivers = sorted({str(r.get("UserName", "")).strip() for r in all_excel if r.get("UserName")})
    carriers_for_dropdown = ["Vodafone", "Orange", "Etisalat", "We"]

    return render_template(
        "trips.html",
        driver_filter=driver_filter,
        trips=page_data,
        trip_id_search=trip_id_search,
        route_quality_filter=route_quality_filter,
        model_filter=model_filter,
        ram_filter=ram_filter,
        carrier_filter=carrier_filter,
        variance_min=variance_min if variance_min is not None else "",
        variance_max=variance_max if variance_max is not None else "",
        trip_time=trip_time_filter,
        trip_time_op=trip_time_op,
        completed_by=completed_by_filter,
        log_count=log_count_filter,
        log_count_op=log_count_op,
        status=status_filter,
        lack_of_accuracy_filter=lack_of_accuracy_filter,
        tags_filter=tags_filter,
        total_rows=total_rows,
        page=page,
        total_pages=total_pages,
        page_size=page_size,
        min_date=min_date,
        max_date=max_date,
        drivers=drivers,
        carriers_for_dropdown=carriers_for_dropdown,
        statuses=statuses,
        completed_by_options=completed_by_options,
        models_options=models_options,
        tags_for_dropdown=tags_for_dropdown,
        expected_trip_quality_filter=expected_trip_quality_filter,
        filters=filters  # Pass all active filters to the template
    )






@app.route("/trip/<int:trip_id>")
def trip_detail(trip_id):
    """
    Show detail page for a single trip, merges with DB.
    """
    session_local = db_session()
    db_trip, update_status = update_trip_db(trip_id)
    
    # Ensure update_status has all required keys even if there was an error
    if "error" in update_status:
        update_status = {
            "needed_update": False,
            "record_exists": True if db_trip else False,
            "updated_fields": [],
            "reason_for_update": ["Error: " + update_status.get("error", "Unknown error")],
            "error": update_status["error"]
        }
    

    if db_trip and db_trip.status and db_trip.status.lower() == "completed":
        api_data = None
    else:
        api_data = fetch_trip_from_api(trip_id)
    trip_attributes = {}
    if api_data and "data" in api_data:
        trip_attributes = api_data["data"]["attributes"]

    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)
    excel_trip_data = None
    for row in excel_data:
        if row.get("tripId") == trip_id:
            excel_trip_data = row
            break

    distance_verification = "N/A"
    trip_insight = ""
    distance_percentage = "N/A"
    if db_trip:
        try:
            md = float(db_trip.manual_distance)
        except (TypeError, ValueError):
            md = None
        try:
            cd = float(db_trip.calculated_distance)
        except (TypeError, ValueError):
            cd = None
        if md is not None and cd is not None:
            lower_bound = md * 0.8
            upper_bound = md * 1.2
            if lower_bound <= cd <= upper_bound:
                distance_verification = "Calculated distance is true"
                trip_insight = "Trip data is consistent."
            else:
                distance_verification = "Manual distance is true"
                trip_insight = "Trip data is inconsistent."
            if md != 0:
                distance_percentage = f"{(cd / md * 100):.2f}%"
        else:
            distance_verification = "N/A"
            trip_insight = "N/A"
            distance_percentage = "N/A"

    session_local.close()
    return render_template(
        "trip_detail.html",
        db_trip=db_trip,
        trip_attributes=trip_attributes,
        excel_trip_data=excel_trip_data,
        distance_verification=distance_verification,
        trip_insight=trip_insight,
        distance_percentage=distance_percentage,
        update_status=update_status
    )

@app.route("/update_route_quality", methods=["POST"])
def update_route_quality():
    """
    AJAX endpoint to update route_quality for a given trip_id.
    """
    session_local = db_session()
    data = request.get_json()
    trip_id = data.get("trip_id")
    quality = data.get("route_quality")
    db_trip = session_local.query(Trip).filter_by(trip_id=trip_id).first()
    if not db_trip:
        db_trip = Trip(
            trip_id=trip_id,
            route_quality=quality,
            status="",
            manual_distance=None,
            calculated_distance=None
        )
        session_local.add(db_trip)
    else:
        db_trip.route_quality = quality
    session_local.commit()
    session_local.close()
    return jsonify({"status": "success", "message": "Route quality updated."}), 200

@app.route("/update_trip_tags", methods=["POST"])
def update_trip_tags():
    session_local = db_session()
    data = request.get_json()
    trip_id = data.get("trip_id")
    tags_list = data.get("tags", [])
    if not trip_id:
        session_local.close()
        return jsonify({"status": "error", "message": "trip_id is required"}), 400
    trip = session_local.query(Trip).filter_by(trip_id=trip_id).first()
    if not trip:
        session_local.close()
        return jsonify({"status": "error", "message": "Trip not found"}), 404
    # Clear existing tags
    trip.tags = []
    updated_tags = []
    for tag_name in tags_list:
        tag = session_local.query(Tag).filter_by(name=tag_name).first()
        if not tag:
            tag = Tag(name=tag_name)
            session_local.add(tag)
            session_local.flush()
        trip.tags.append(tag)
        updated_tags.append(tag.name)
    session_local.commit()
    session_local.close()
    return jsonify({"status": "success", "tags": updated_tags}), 200

@app.route("/get_tags", methods=["GET"])
def get_tags():
    session_local = db_session()
    tags = session_local.query(Tag).all()
    data = [{"id": tag.id, "name": tag.name} for tag in tags]
    session_local.close()
    return jsonify({"status": "success", "tags": data}), 200

@app.route("/create_tag", methods=["POST"])
def create_tag():
    session_local = db_session()
    data = request.get_json()
    tag_name = data.get("name")
    if not tag_name:
        session_local.close()
        return jsonify({"status": "error", "message": "Tag name is required"}), 400
    existing = session_local.query(Tag).filter_by(name=tag_name).first()
    if existing:
        session_local.close()
        return jsonify({"status": "error", "message": "Tag already exists"}), 400
    tag = Tag(name=tag_name)
    session_local.add(tag)
    session_local.commit()
    session_local.refresh(tag)
    session_local.close()
    return jsonify({"status": "success", "tag": {"id": tag.id, "name": tag.name}}), 200

@app.route("/trip_insights")
def trip_insights():
    """
    Shows route quality counts, distance averages, distance consistency, and additional dashboards:
      - Average Trip Duration vs Trip Quality
      - Completed By vs Trip Quality
      - Average Logs Count vs Trip Quality
      - App Version vs Trip Quality

    Now uses a new query parameter quality_metric which can be:
      "manual"   -> use manual quality (statuses: No Logs Trips, Trip Points Only Exist, Low, Moderate, High)
      "expected" -> use expected quality (statuses: No Logs Trip, Trip Points Only Exist, Low Quality Trip, Moderate Quality Trip, High Quality Trip)
    """
    from datetime import datetime
    from collections import defaultdict, Counter

    session_local = db_session()
    data_scope = flask_session.get("data_scope", "all")

    # Load Excel data and get trip IDs
    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)
    excel_trip_ids = [r["tripId"] for r in excel_data if r.get("tripId")]

    if data_scope == "excel":
        trips_db = session_local.query(Trip).filter(Trip.trip_id.in_(excel_trip_ids)).all()
    else:
        trips_db = session_local.query(Trip).all()

    # Use manual quality from route_quality field
    quality_metric = "manual"
    possible_statuses = ["No Logs Trips", "Trip Points Only Exist", "Low", "Moderate", "High"]
    quality_counts = {status: 0 for status in possible_statuses}
    quality_counts[""] = 0

    total_manual = 0
    total_calculated = 0
    count_manual = 0
    count_calculated = 0
    consistent = 0
    inconsistent = 0

    # Aggregation: loop over trips and use the selected quality value
    for trip in trips_db:
        quality = trip.route_quality if trip.route_quality is not None else ""
        quality = quality.strip() if isinstance(quality, str) else ""
        if quality in quality_counts:
            quality_counts[quality] += 1
        else:
            quality_counts[""] += 1

        try:
            md = float(trip.manual_distance)
            cd = float(trip.calculated_distance)
            total_manual += md
            total_calculated += cd
            count_manual += 1
            count_calculated += 1
            if md != 0 and abs(cd - md) / md <= 0.2:
                consistent += 1
            else:
                inconsistent += 1
        except:
            pass

    avg_manual = total_manual / count_manual if count_manual else 0
    avg_calculated = total_calculated / count_calculated if count_calculated else 0

    # Build excel_map from Excel data
    excel_map = {r['tripId']: r for r in excel_data if r.get('tripId')}

    # Device specs aggregation using manual quality
    device_specs = defaultdict(lambda: defaultdict(list))
    for trip in trips_db:
        trip_id = trip.trip_id
        quality = trip.route_quality if trip.route_quality is not None else "Unknown"
        quality = quality.strip() if isinstance(quality, str) else "Unknown"
        if trip_id in excel_map:
            row = excel_map[trip_id]
            device_specs[quality]['model'].append(row.get('model', 'Unknown'))
            device_specs[quality]['android'].append(row.get('Android Version', 'Unknown'))
            device_specs[quality]['manufacturer'].append(row.get('manufacturer', 'Unknown'))
            device_specs[quality]['ram'].append(row.get('RAM', 'Unknown'))

    # Build insights text based on manual quality
    manual_insights = {}
    for quality, specs in device_specs.items():
        model_counter = Counter(specs['model'])
        android_counter = Counter(specs['android'])
        manufacturer_counter = Counter(specs['manufacturer'])
        ram_counter = Counter(specs['ram'])
        most_common_model = model_counter.most_common(1)[0][0] if model_counter else 'N/A'
        most_common_android = android_counter.most_common(1)[0][0] if android_counter else 'N/A'
        most_common_manufacturer = manufacturer_counter.most_common(1)[0][0] if manufacturer_counter else 'N/A'
        most_common_ram = ram_counter.most_common(1)[0][0] if ram_counter else 'N/A'
        insight = f"For trips with quality '{quality}', most devices are {most_common_manufacturer} {most_common_model} (Android {most_common_android}, RAM {most_common_ram})."
        if quality.lower() == "high":
            insight += " This suggests that high quality trips are associated with robust mobile specs, contributing to accurate tracking."
        elif quality.lower() == "low":
            insight += " This might indicate that lower quality trips could be influenced by devices with suboptimal specifications."
        manual_insights[quality] = insight

    # Aggregation: Lack of Accuracy vs Manual Trip Quality
    accuracy_data = {}
    for trip in trips_db:
        quality = trip.route_quality if trip.route_quality is not None else "Unspecified"
        quality = quality.strip() if isinstance(quality, str) else "Unspecified"
        if quality not in accuracy_data:
            accuracy_data[quality] = {"count": 0, "lack_count": 0}
        accuracy_data[quality]["count"] += 1
        if trip.lack_of_accuracy:
            accuracy_data[quality]["lack_count"] += 1
    accuracy_percentages = {}
    for quality, data in accuracy_data.items():
        count = data["count"]
        lack = data["lack_count"]
        percentage = round((lack / count) * 100, 2) if count > 0 else 0
        accuracy_percentages[quality] = percentage

    # Dashboard Aggregations based on manual quality

    # 1. Average Trip Duration vs Manual Trip Quality
    trip_duration_sum = {}
    trip_duration_count = {}
    for trip in trips_db:
        quality = trip.route_quality if trip.route_quality is not None else "Unspecified"

        quality = quality.strip() if isinstance(quality, str) else "Unspecified"
        if trip.trip_time is not None and trip.trip_time != "":
            trip_duration_sum[quality] = trip_duration_sum.get(quality, 0) + float(trip.trip_time)
            trip_duration_count[quality] = trip_duration_count.get(quality, 0) + 1
    avg_trip_duration_quality = {}
    for quality in trip_duration_sum:
        avg_trip_duration_quality[quality] = trip_duration_sum[quality] / trip_duration_count[quality]

    # 2. Completed By vs Manual Trip Quality
    completed_by_quality = {}
    for trip in trips_db:
        quality = trip.route_quality if trip.route_quality is not None else "Unspecified"
        quality = quality.strip() if isinstance(quality, str) else "Unspecified"
        comp = trip.completed_by if trip.completed_by else "Unknown"
        if quality not in completed_by_quality:
            completed_by_quality[quality] = {}
        completed_by_quality[quality][comp] = completed_by_quality[quality].get(comp, 0) + 1

    # 3. Average Logs Count vs Manual Trip Quality
    logs_sum = {}
    logs_count = {}
    for trip in trips_db:
        quality = trip.route_quality if trip.route_quality is not None else "Unspecified"
        quality = quality.strip() if isinstance(quality, str) else "Unspecified"
        if trip.coordinate_count is not None and trip.coordinate_count != "":
            logs_sum[quality] = logs_sum.get(quality, 0) + int(trip.coordinate_count)
            logs_count[quality] = logs_count.get(quality, 0) + 1
    avg_logs_count_quality = {}
    for quality in logs_sum:
        avg_logs_count_quality[quality] = logs_sum[quality] / logs_count[quality]

    # 4. App Version vs Manual Trip Quality
    app_version_quality = {}
    for trip in trips_db:
        row = excel_map.get(trip.trip_id)
        if row:
            app_ver = row.get("app_version", "Unknown")
        else:
            app_ver = "Unknown"
        quality = trip.route_quality if trip.route_quality is not None else "Unspecified"
        quality = quality.strip() if isinstance(quality, str) else "Unspecified"
        if app_ver not in app_version_quality:
            app_version_quality[app_ver] = {}
        app_version_quality[app_ver][quality] = app_version_quality[app_ver].get(quality, 0) + 1

    # Additional Aggregations for manual quality


    quality_drilldown = {}
    for trip in trips_db:
        if quality_metric == "expected":
            quality = trip.expected_trip_quality if trip.expected_trip_quality is not None else "Unspecified"
        else:
            quality = trip.route_quality if trip.route_quality is not None else "Unspecified"
        quality = quality.strip() if isinstance(quality, str) else "Unspecified"
        # Build the device specs based on quality; using our previously built device_specs dict is sufficient.
        # (We assume device_specs keys already reflect the chosen quality as built above.)
    # We'll assume quality_drilldown is built based on device_specs dict keys.
    for quality, specs in device_specs.items():
        quality_drilldown[quality] = {
            'model': dict(Counter(specs['model'])),
            'android': dict(Counter(specs['android'])),
            'manufacturer': dict(Counter(specs['manufacturer'])),
            'ram': dict(Counter(specs['ram']))
        }

    allowed_ram_str = ["2GB", "3GB", "4GB", "6GB", "8GB", "12GB", "16GB"]
    ram_quality_counts = {ram: {} for ram in allowed_ram_str}
    import re
    for trip in trips_db:
        quality_val = trip.route_quality if trip.route_quality is not None else "Unspecified"
        quality_val = quality_val.strip() if isinstance(quality_val, str) else "Unspecified"
        row = excel_map.get(trip.trip_id)
        if row:
            ram_str = row.get("RAM", "")
            match = re.search(r'(\d+(?:\.\d+)?)', str(ram_str))
            if match:
                ram_value = float(match.group(1))
                try:
                    ram_int = int(round(ram_value))
                except:
                    continue
                nearest = min([2, 3, 4, 6, 8, 12, 16], key=lambda v: abs(v - ram_int))
                ram_label = f"{nearest}GB"
                if quality_val not in ["High", "Moderate", "Low", "No Logs Trips", "Trip Points Only Exist"]:
                    quality_val = "Empty"
                if quality_val not in ram_quality_counts[ram_label]:
                    ram_quality_counts[ram_label][quality_val] = 0
                ram_quality_counts[ram_label][quality_val] += 1

    sensor_cols = ["Fingerprint Sensor", "Accelerometer", "Gyro",
                   "Proximity Sensor", "Compass", "Barometer",
                   "Background Task Killing Tendency"]
    sensor_stats = {}
    for sensor in sensor_cols:
        sensor_stats[sensor] = {}
    for trip in trips_db:
        quality_val = trip.route_quality if trip.route_quality is not None else "Unspecified"
        quality_val = quality_val.strip() if isinstance(quality_val, str) else "Unspecified"
        row = excel_map.get(trip.trip_id)
        if row:
            for sensor in sensor_cols:
                value = row.get(sensor, "")
                present = False
                if isinstance(value, str) and value.lower() == "true":
                    present = True
                elif value is True:
                    present = True
                if quality_val not in sensor_stats[sensor]:
                    sensor_stats[sensor][quality_val] = {"present": 0, "total": 0}
                sensor_stats[sensor][quality_val]["total"] += 1
                if present:
                    sensor_stats[sensor][quality_val]["present"] += 1

    quality_by_os = {}
    for trip in trips_db:
        row = excel_map.get(trip.trip_id)
        if row:
            os_ver = row.get("Android Version", "Unknown")
            q = trip.route_quality if trip.route_quality is not None else "Unspecified"
            q = q.strip() if isinstance(q, str) else "Unspecified"
            if os_ver not in quality_by_os:
                quality_by_os[os_ver] = {}
            quality_by_os[os_ver][q] = quality_by_os[os_ver].get(q, 0) + 1

    manufacturer_quality = {}
    for trip in trips_db:
        row = excel_map.get(trip.trip_id)
        if row:
            manu = row.get("manufacturer", "Unknown")
            q = trip.route_quality if trip.route_quality is not None else "Unspecified"
            q = q.strip() if isinstance(q, str) else "Unspecified"
            if manu not in manufacturer_quality:
                manufacturer_quality[manu] = {}
            manufacturer_quality[manu][q] = manufacturer_quality[manu].get(q, 0) + 1

    carrier_quality = {}
    for trip in trips_db:
        row = excel_map.get(trip.trip_id)
        if row:
            carrier_val = normalize_carrier(row.get("carrier", "Unknown"))
            q = trip.route_quality if trip.route_quality is not None else "Unspecified"

            q = q.strip() if isinstance(q, str) else "Unspecified"
            if carrier_val not in carrier_quality:
                carrier_quality[carrier_val] = {}
            carrier_quality[carrier_val][q] = carrier_quality[carrier_val].get(q, 0) + 1

    time_series = {}
    for row in excel_data:
        try:
            time_str = row.get("time", "")
            if time_str:
                dt = datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
                date_str = dt.strftime("%Y-%m-%d")
                # For time series, we use the manual quality from Excel data (assuming it's stored in "route_quality")
                q = row.get("route_quality", "Unspecified")
                if date_str not in time_series:
                    time_series[date_str] = {}
                time_series[date_str][q] = time_series[date_str].get(q, 0) + 1
        except:
            continue

    session_local.close()
    return render_template(
        "trip_insights.html",
        quality_counts=quality_counts,
        avg_manual=avg_manual,
        avg_calculated=avg_calculated,
        consistent=consistent,
        inconsistent=inconsistent,
        automatic_insights=manual_insights,
        quality_drilldown=quality_drilldown,
        ram_quality_counts=ram_quality_counts,
        sensor_stats=sensor_stats,
        quality_by_os=quality_by_os,
        manufacturer_quality=manufacturer_quality,
        carrier_quality=carrier_quality,
        time_series=time_series,
        avg_trip_duration_quality=avg_trip_duration_quality,
        completed_by_quality=completed_by_quality,
        avg_logs_count_quality=avg_logs_count_quality,
        app_version_quality=app_version_quality,
        accuracy_data=accuracy_percentages,
        quality_metric="manual"
    )


@app.route("/automatic_insights")
def automatic_insights():
    """
    Shows trip insights based on the expected trip quality (automatic),
    including:
      - Filtering out trips with calculated_distance > 2000 km.
      - Calculating average distance variance, accurate counts, etc.
      - Handling trip_time outliers and possible seconds→hours conversion.
      - Building a time_series from the same filtered trips for the temporal trends chart.
      - Trying multiple date/time formats to parse 'time' from Excel data so the chart won't be empty.
    """
    from datetime import datetime
    from collections import defaultdict, Counter
    import re

    session_local = db_session()
    data_scope = flask_session.get("data_scope", "all")

    # 1) Load Excel data and build a tripId→Excel row mapping
    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)
    excel_map = {r['tripId']: r for r in excel_data if r.get('tripId')}
    excel_trip_ids = list(excel_map.keys())

    # 2) Query DB trips, optionally restricting to those in Excel
    if data_scope == "excel":
        trips_db = session_local.query(Trip).filter(Trip.trip_id.in_(excel_trip_ids)).all()
    else:
        trips_db = session_local.query(Trip).all()

    # 3) Filter out trips with calc_distance > 2000 km
    filtered_trips = []
    for trip in trips_db:
        try:
            cd = float(trip.calculated_distance)
            if cd <= 2000:
                filtered_trips.append(trip)
        except:
            continue

    # 4) Initialize metrics
    possible_statuses = [
        "No Logs Trip", 
        "Trip Points Only Exist", 
        "Low Quality Trip", 
        "Moderate Quality Trip", 
        "High Quality Trip"
    ]
    quality_counts = {status: 0 for status in possible_statuses}
    quality_counts[""] = 0

    total_manual = 0.0
    total_calculated = 0.0
    count_manual = 0
    count_calculated = 0
    consistent = 0
    inconsistent = 0

    variance_sum = 0.0
    variance_count = 0
    accurate_count = 0
    app_killed_count = 0
    one_log_count = 0
    total_short_dist = 0.0
    total_medium_dist = 0.0
    total_long_dist = 0.0

    driver_totals = defaultdict(int)       # driver_name → total trips
    driver_counts = defaultdict(lambda: defaultdict(int))  # driver_name → {quality: count}

    # 5) Main loop: gather metrics from filtered trips
    for trip in filtered_trips:
        eq_quality = (trip.expected_trip_quality or "").strip()
        if eq_quality in quality_counts:
            quality_counts[eq_quality] += 1
        else:
            quality_counts[""] += 1

        # Distances & variance
        try:
            md = float(trip.manual_distance)
            cd = float(trip.calculated_distance)
            total_manual += md
            total_calculated += cd
            count_manual += 1
            count_calculated += 1

            if md > 0:
                variance = abs(cd - md) / md * 100
                variance_sum += variance
                variance_count += 1
                if variance < 25.0:
                    accurate_count += 1

            if md > 0 and abs(cd - md) / md <= 0.2:
                consistent += 1
            else:
                inconsistent += 1
        except:
            pass

        # Summation of short/medium/long
        if trip.short_segments_distance:
            total_short_dist += float(trip.short_segments_distance)
        if trip.medium_segments_distance:
            total_medium_dist += float(trip.medium_segments_distance)
        if trip.long_segments_distance:
            total_long_dist += float(trip.long_segments_distance)

        # Single-log trips
        if trip.coordinate_count == 1:
            one_log_count += 1

        # "App killed" issue
        try:
            if trip.lack_of_accuracy is False and float(trip.calculated_distance) > 0:
                lm_distance = (float(trip.medium_segments_distance or 0) 
                               + float(trip.long_segments_distance or 0))
                lm_count = (trip.medium_segments_count or 0) + (trip.long_segments_count or 0)
                if lm_count > 0 and (lm_distance / float(trip.calculated_distance)) >= 0.4:
                    app_killed_count += 1
        except:
            pass

        # Driver name
        driver_name = getattr(trip, 'driver_name', None)
        if not driver_name and trip.trip_id in excel_map:
            driver_name = excel_map[trip.trip_id].get("UserName")
        if driver_name:
            driver_totals[driver_name] += 1
            driver_counts[driver_name][eq_quality] += 1

    # 6) Final aggregates
    avg_manual = total_manual / count_manual if count_manual else 0
    avg_calculated = total_calculated / count_calculated if count_calculated else 0
    avg_distance_variance = variance_sum / variance_count if variance_count else 0
    total_trips = len(filtered_trips)

    accurate_count_pct = (accurate_count / total_trips * 100) if total_trips else 0
    app_killed_pct = (app_killed_count / total_trips * 100) if total_trips else 0
    one_log_pct = (one_log_count / total_trips * 100) if total_trips else 0
    short_dist_pct = (total_short_dist / total_calculated * 100) if total_calculated else 0
    medium_dist_pct = (total_medium_dist / total_calculated * 100) if total_calculated else 0
    long_dist_pct = (total_long_dist / total_calculated * 100) if total_calculated else 0

    # 7) Average Trip Duration vs Expected Quality
    trip_duration_sum = {}
    trip_duration_count = {}
    for trip in filtered_trips:
        q = (trip.expected_trip_quality or "Unspecified").strip()
        try:
            raw_tt = float(trip.trip_time)
        except:
            continue

        # Convert possible seconds→hours if over 72
        if raw_tt > 72:
            raw_tt /= 3600.0
        # Skip if >720 hours or negative
        if raw_tt < 0 or raw_tt > 720:
            continue

        trip_duration_sum[q] = trip_duration_sum.get(q, 0) + raw_tt
        trip_duration_count[q] = trip_duration_count.get(q, 0) + 1

    avg_trip_duration_quality = {}
    for q in trip_duration_sum:
        c = trip_duration_count[q]
        if c > 0:
            avg_trip_duration_quality[q] = trip_duration_sum[q] / c

    # 8) Build device specs & additional charts from filtered trips
    device_specs = defaultdict(lambda: defaultdict(list))
    for trip in filtered_trips:
        q = (trip.expected_trip_quality or "Unknown").strip()
        row = excel_map.get(trip.trip_id)
        if not row:
            continue
        device_specs[q]['model'].append(row.get('model','Unknown'))
        device_specs[q]['android'].append(row.get('Android Version','Unknown'))
        device_specs[q]['manufacturer'].append(row.get('manufacturer','Unknown'))
        device_specs[q]['ram'].append(row.get('RAM','Unknown'))

    # Generate a text insight for each quality
    automatic_insights_text = {}
    for quality, specs in device_specs.items():
        model_counter = Counter(specs['model'])
        android_counter = Counter(specs['android'])
        manu_counter = Counter(specs['manufacturer'])
        ram_counter = Counter(specs['ram'])

        mc_model = model_counter.most_common(1)[0][0] if model_counter else 'N/A'
        mc_android = android_counter.most_common(1)[0][0] if android_counter else 'N/A'
        mc_manu = manu_counter.most_common(1)[0][0] if manu_counter else 'N/A'
        mc_ram = ram_counter.most_common(1)[0][0] if ram_counter else 'N/A'
        insight = f"For '{quality}', common device is {mc_manu} {mc_model} (Android {mc_android}, RAM {mc_ram})."
        if quality.lower() == 'high quality trip':
            insight += " Suggests better specs correlate with high quality."
        elif quality.lower() == 'low quality trip':
            insight += " Possibly indicates suboptimal specs or usage."
        automatic_insights_text[quality] = insight

    # 9) Lack of Accuracy vs Expected Trip Quality
    accuracy_data = {}
    for trip in filtered_trips:
        q = (trip.expected_trip_quality or "Unspecified").strip()
        if q not in accuracy_data:
            accuracy_data[q] = {"count":0,"lack_count":0}
        accuracy_data[q]["count"] += 1
        if trip.lack_of_accuracy:
            accuracy_data[q]["lack_count"] += 1

    accuracy_percentages = {}
    for q, d in accuracy_data.items():
        if d["count"]>0:
            accuracy_percentages[q] = round((d["lack_count"]/d["count"])*100,2)
        else:
            accuracy_percentages[q] = 0

    # 10) Completed By vs Expected Quality
    completed_by_quality = {}
    for trip in filtered_trips:
        q = (trip.expected_trip_quality or "Unspecified").strip()
        comp = trip.completed_by if trip.completed_by else "Unknown"
        if q not in completed_by_quality:
            completed_by_quality[q] = {}
        completed_by_quality[q][comp] = completed_by_quality[q].get(comp,0)+1

    # 11) Average Logs Count vs Expected Quality
    logs_sum = {}
    logs_count = {}
    for trip in filtered_trips:
        q = (trip.expected_trip_quality or "Unspecified").strip()
        if trip.coordinate_count:
            logs_sum[q] = logs_sum.get(q,0)+trip.coordinate_count
            logs_count[q] = logs_count.get(q,0)+1
    avg_logs_count_quality = {}
    for q in logs_sum:
        if logs_count[q]>0:
            avg_logs_count_quality[q] = logs_sum[q]/logs_count[q]

    # 12) App Version vs Expected Quality
    app_version_quality = {}
    for trip in filtered_trips:
        row = excel_map.get(trip.trip_id)
        if not row:
            continue
        ver = row.get("app_version","Unknown")
        q = (trip.expected_trip_quality or "Unspecified").strip()
        if ver not in app_version_quality:
            app_version_quality[ver] = {}
        app_version_quality[ver][q] = app_version_quality[ver].get(q,0)+1

    # 13) Quality Drilldown
    quality_drilldown = {}
    for q, specs in device_specs.items():
        quality_drilldown[q] = {
            'model': dict(Counter(specs['model'])),
            'android': dict(Counter(specs['android'])),
            'manufacturer': dict(Counter(specs['manufacturer'])),
            'ram': dict(Counter(specs['ram']))
        }

    # 14) RAM Quality Aggregation
    allowed_ram_str = ["2GB","3GB","4GB","6GB","8GB","12GB","16GB"]
    ram_quality_counts = {ram:{} for ram in allowed_ram_str}
    for trip in filtered_trips:
        row = excel_map.get(trip.trip_id)
        if not row:
            continue
        q = (trip.expected_trip_quality or "Unspecified").strip()
        ram_str = row.get("RAM","")
        m = re.search(r'(\d+(?:\.\d+)?)', str(ram_str))
        if not m:
            continue
        try:
            val = float(m.group(1))
            val_int = int(round(val))
        except:
            continue
        nearest = min([2,3,4,6,8,12,16], key=lambda v: abs(v - val_int))
        label = f"{nearest}GB"
        if q not in ["High Quality Trip","Moderate Quality Trip","Low Quality Trip","No Logs Trip","Trip Points Only Exist"]:
            q = "Empty"
        ram_quality_counts[label][q] = ram_quality_counts[label].get(q,0)+1

    # 15) Sensor & Feature Aggregation
    sensor_cols = ["Fingerprint Sensor","Accelerometer","Gyro",
                   "Proximity Sensor","Compass","Barometer",
                   "Background Task Killing Tendency"]
    sensor_stats = {s:{} for s in sensor_cols}
    for trip in filtered_trips:
        row = excel_map.get(trip.trip_id)
        if not row:
            continue
        q = (trip.expected_trip_quality or "Unspecified").strip()
        for s in sensor_cols:
            val = row.get(s,"")
            present = ((isinstance(val,str) and val.lower()=="true") or val is True)
            if q not in sensor_stats[s]:
                sensor_stats[s][q] = {"present":0,"total":0}
            sensor_stats[s][q]["total"] += 1
            if present:
                sensor_stats[s][q]["present"] += 1

    # 16) Quality by OS
    quality_by_os = {}
    for trip in filtered_trips:
        row = excel_map.get(trip.trip_id)
        if not row:
            continue
        os_ver = row.get("Android Version","Unknown")
        q = (trip.expected_trip_quality or "Unspecified").strip()
        if os_ver not in quality_by_os:
            quality_by_os[os_ver] = {}
        quality_by_os[os_ver][q] = quality_by_os[os_ver].get(q,0)+1

    # 17) Manufacturer Quality
    manufacturer_quality = {}
    for trip in filtered_trips:
        row = excel_map.get(trip.trip_id)
        if not row:
            continue
        manu = row.get("manufacturer","Unknown")
        q = (trip.expected_trip_quality or "Unspecified").strip()
        if manu not in manufacturer_quality:
            manufacturer_quality[manu] = {}
        manufacturer_quality[manu][q] = manufacturer_quality[manu].get(q,0)+1

    # 18) Carrier Quality
    carrier_quality = {}
    for trip in filtered_trips:
        row = excel_map.get(trip.trip_id)
        if not row:
            continue
        cval = normalize_carrier(row.get("carrier","Unknown"))
        q = (trip.expected_trip_quality or "Unspecified").strip()
        if cval not in carrier_quality:
            carrier_quality[cval] = {}
        carrier_quality[cval][q] = carrier_quality[cval].get(q,0)+1

    # --------------------- FIXING THE TIME SERIES ---------------------
    # We'll parse 'time' from the same filtered trips & attempt multiple formats
    # so the chart has consistent data.
    POSSIBLE_TIME_FORMATS = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y %H:%M:%S"
    ]

    time_series = {}
    for trip in filtered_trips:
        row = excel_map.get(trip.trip_id)
        if not row:
            continue
        time_str = row.get("time", "")
        if not time_str:
            continue

        dt_obj = None
        for fmt in POSSIBLE_TIME_FORMATS:
            try:
                dt_obj = datetime.strptime(time_str, fmt)
                break
            except:
                pass
        if not dt_obj:
            # Could not parse date in known formats
            continue

        date_str = dt_obj.strftime("%Y-%m-%d")
        eq = (trip.expected_trip_quality or "Unspecified").strip()
        if date_str not in time_series:
            time_series[date_str] = {}
        time_series[date_str][eq] = time_series[date_str].get(eq, 0) + 1

    # 19) Driver Behavior Analysis with threshold ratio
    threshold = 0.75
    top_high_drivers = []
    top_moderate_drivers = []
    top_low_drivers = []
    top_no_logs_drivers = []
    top_points_only_drivers = []

    for driver, total in driver_totals.items():
        if total <= 0:
            continue
        ratio_high = driver_counts[driver].get("High Quality Trip",0)/total
        ratio_mod = driver_counts[driver].get("Moderate Quality Trip",0)/total
        ratio_low = driver_counts[driver].get("Low Quality Trip",0)/total
        ratio_no_logs = driver_counts[driver].get("No Logs Trip",0)/total
        ratio_points = driver_counts[driver].get("Trip Points Only Exist",0)/total

        if ratio_high >= threshold:
            top_high_drivers.append((driver, ratio_high))
        if ratio_mod >= threshold:
            top_moderate_drivers.append((driver, ratio_mod))
        if ratio_low >= threshold:
            top_low_drivers.append((driver, ratio_low))
        if ratio_no_logs >= threshold:
            top_no_logs_drivers.append((driver, ratio_no_logs))
        if ratio_points >= threshold:
            top_points_only_drivers.append((driver, ratio_points))

    # Sort each driver group by ratio desc, pick top 3
    top_high_drivers = [d for d,r in sorted(top_high_drivers,key=lambda x:x[1],reverse=True)[:3]]
    top_moderate_drivers = [d for d,r in sorted(top_moderate_drivers,key=lambda x:x[1],reverse=True)[:3]]
    top_low_drivers = [d for d,r in sorted(top_low_drivers,key=lambda x:x[1],reverse=True)[:3]]
    top_no_logs_drivers = [d for d,r in sorted(top_no_logs_drivers,key=lambda x:x[1],reverse=True)[:3]]
    top_points_only_drivers = [d for d,r in sorted(top_points_only_drivers,key=lambda x:x[1],reverse=True)[:3]]

    session_local.close()

    return render_template(
        "Automatic_insights.html",
        # Basic quality counts
        quality_counts=quality_counts,
        # Distances & variance
        avg_manual=avg_manual,
        avg_calculated=avg_calculated,
        consistent=consistent,
        inconsistent=inconsistent,
        avg_distance_variance=avg_distance_variance,
        accurate_count=accurate_count,
        accurate_count_pct=accurate_count_pct,
        app_killed_count=app_killed_count,
        app_killed_pct=app_killed_pct,
        one_log_count=one_log_count,
        one_log_pct=one_log_pct,
        short_dist_pct=short_dist_pct,
        medium_dist_pct=medium_dist_pct,
        long_dist_pct=long_dist_pct,

        # Duration, logs, versions, etc.
        avg_trip_duration_quality=avg_trip_duration_quality,
        completed_by_quality=completed_by_quality,
        avg_logs_count_quality=avg_logs_count_quality,
        app_version_quality=app_version_quality,

        # Additional data for charts
        automatic_insights=automatic_insights_text,
        quality_drilldown=quality_drilldown,
        ram_quality_counts=ram_quality_counts,
        sensor_stats=sensor_stats,
        quality_by_os=quality_by_os,
        manufacturer_quality=manufacturer_quality,
        carrier_quality=carrier_quality,

        # The fixed time_series with multi-format parsing
        time_series=time_series,

        # Accuracy data
        accuracy_data=accuracy_percentages,
        quality_metric="expected",

        # Driver Behavior
        top_high_drivers=top_high_drivers,
        top_moderate_drivers=top_moderate_drivers,
        top_low_drivers=top_low_drivers,
        top_no_logs_drivers=top_no_logs_drivers,
        top_points_only_drivers=top_points_only_drivers
    )










@app.route("/save_filter", methods=["POST"])
def save_filter():
    """
    Store current filter parameters in session under a filter name.
    """
    filter_name = request.form.get("filter_name")
    filters = {
        "trip_id": request.form.get("trip_id"),
        "route_quality": request.form.get("route_quality"),
        "model": request.form.get("model"),
        "ram": request.form.get("ram"),
        "carrier": request.form.get("carrier"),
        "variance_min": request.form.get("variance_min"),
        "variance_max": request.form.get("variance_max"),
        "driver": request.form.get("driver")
    }
    if filter_name:
        saved = flask_session.get("saved_filters", {})
        saved[filter_name] = filters
        flask_session["saved_filters"] = saved
        flash(f"Filter '{filter_name}' saved.", "success")
    else:
        flash("Please provide a filter name.", "danger")
    return redirect(url_for("trips"))

@app.route("/apply_filter/<filter_name>")
def apply_filter(filter_name):
    """
    Apply a saved filter by redirecting to /trips with the saved query params.
    """
    saved = flask_session.get("saved_filters", {})
    filters = saved.get(filter_name)
    if filters:
        qs = "&".join(f"{key}={value}" for key, value in filters.items() if value)
        return redirect(url_for("trips") + "?" + qs)
    else:
        flash("Saved filter not found.", "danger")
        return redirect(url_for("trips"))

@app.route('/update_date_range', methods=['POST'])
def update_date_range():
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    if not start_date or not end_date:
        return jsonify({'error': 'Both start_date and end_date are required.'}), 400

    # Backup existing consolidated data
    data_file = 'data/data.xlsx'
    backup_dir = 'data/backup'
    if os.path.exists(data_file):
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        backup_file = os.path.join(backup_dir, f"data_{start_date}_{end_date}.xlsx")
        try:
            shutil.move(data_file, backup_file)
        except Exception as e:
            return jsonify({'error': 'Failed to backup data file: ' + str(e)}), 500

    # Run exportmix.py with new dates
    try:
        # Fix: Call with the correct command-line arguments
        subprocess.check_call(['python3', 'exportmix.py', '--start-date', start_date, '--end-date', end_date])
    except subprocess.CalledProcessError as e:
        return jsonify({'error': 'Failed to export data: ' + str(e)}), 500

    # Run consolidatemixpanel.py
    try:
        subprocess.check_call(['python3', 'consolidatemixpanel.py'])
    except subprocess.CalledProcessError as e:
        return jsonify({'error': 'Failed to consolidate data: ' + str(e)}), 500

    return jsonify({'message': 'Data updated successfully.'})

@app.route("/update_db_async", methods=["POST"])
def update_db_async():
    job_id = str(uuid.uuid4())
    update_jobs[job_id] = {
        "status": "processing", 
        "total": 0, 
        "completed": 0, 
        "updated": 0,
        "skipped": 0,
        "errors": 0, 
        "created": 0,
        "updated_fields": Counter(),
        "reasons": Counter()
    }
    threading.Thread(target=process_update_db_async, args=(job_id,)).start()
    return jsonify({"job_id": job_id})

def process_update_db_async(job_id):
    try:
        excel_path = os.path.join("data", "data.xlsx")
        excel_data = load_excel_data(excel_path)
        trips_to_update = [row.get("tripId") for row in excel_data if row.get("tripId")]
        update_jobs[job_id]["total"] = len(trips_to_update)
        
        # Process trips using ThreadPoolExecutor
        futures_to_trips = {}
        with ThreadPoolExecutor(max_workers=40) as executor:
            # Submit jobs to the executor
            for trip_id in trips_to_update:
                # Use force_update=False to skip complete records
                future = executor.submit(update_trip_db, trip_id, False)
                futures_to_trips[future] = trip_id
            
            # Process results as they complete
            for future in as_completed(futures_to_trips):
                trip_id = futures_to_trips[future]
                try:
                    db_trip, update_status = future.result()
                    
                    # Track statistics similar to update_db route
                    if "error" in update_status:
                        update_jobs[job_id]["errors"] += 1
                    elif not update_status["record_exists"]:
                        update_jobs[job_id]["created"] += 1
                        update_jobs[job_id]["updated"] += 1
                    elif update_status["updated_fields"]:
                        update_jobs[job_id]["updated"] += 1
                        # Count which fields were updated
                        for field in update_status["updated_fields"]:
                            update_jobs[job_id]["updated_fields"][field] = update_jobs[job_id]["updated_fields"].get(field, 0) + 1
                    else:
                        update_jobs[job_id]["skipped"] += 1
                        
                    # Track reasons for updates
                    for reason in update_status.get("reason_for_update", []):
                        update_jobs[job_id]["reasons"][reason] = update_jobs[job_id]["reasons"].get(reason, 0) + 1
                        
                except Exception as e:
                    print(f"Error processing trip {trip_id}: {e}")
                    update_jobs[job_id]["errors"] += 1
                
                update_jobs[job_id]["completed"] += 1
                
        update_jobs[job_id]["status"] = "completed"
        
        # Prepare summary message
        if update_jobs[job_id]["updated"] > 0:
            most_updated_fields = sorted(update_jobs[job_id]["updated_fields"].items(), 
                                         key=lambda x: x[1], reverse=True)[:3]
            update_jobs[job_id]["summary_fields"] = [f"{field} ({count})" for field, count in most_updated_fields]
            
            most_common_reasons = sorted(update_jobs[job_id]["reasons"].items(), 
                                        key=lambda x: x[1], reverse=True)[:3]
            update_jobs[job_id]["summary_reasons"] = [f"{reason} ({count})" for reason, count in most_common_reasons]
        
    except Exception as e:
        update_jobs[job_id]["status"] = "error"
        update_jobs[job_id]["error_message"] = str(e)

@app.route("/update_all_db_async", methods=["POST"])
def update_all_db_async():
    job_id = str(uuid.uuid4())
    update_jobs[job_id] = {
        "status": "processing", 
        "total": 0, 
        "completed": 0, 
        "updated": 0,
        "skipped": 0,
        "errors": 0, 
        "created": 0,
        "updated_fields": Counter(),
        "reasons": Counter()
    }
    threading.Thread(target=process_update_all_db_async, args=(job_id,)).start()
    return jsonify({"job_id": job_id})

def process_update_all_db_async(job_id):
    try:
        # Load trip IDs from Excel instead of getting all trips from DB
        excel_path = os.path.join("data", "data.xlsx")
        excel_data = load_excel_data(excel_path)
        trips_to_update = [row.get("tripId") for row in excel_data if row.get("tripId")]
        update_jobs[job_id]["total"] = len(trips_to_update)
        
        # Process trips using ThreadPoolExecutor
        futures_to_trips = {}
        with ThreadPoolExecutor(max_workers=40) as executor:
            # Submit jobs to the executor
            for trip_id in trips_to_update:
                # Use force_update=True for full update from API
                future = executor.submit(update_trip_db, trip_id, True)
                futures_to_trips[future] = trip_id
            
            # Process results as they complete
            for future in as_completed(futures_to_trips):
                trip_id = futures_to_trips[future]
                try:
                    db_trip, update_status = future.result()
                    
                    # Track statistics
                    if "error" in update_status:
                        update_jobs[job_id]["errors"] += 1
                    elif not update_status["record_exists"]:
                        update_jobs[job_id]["created"] += 1
                        update_jobs[job_id]["updated"] += 1
                    elif update_status["updated_fields"]:
                        update_jobs[job_id]["updated"] += 1
                        # Count which fields were updated
                        for field in update_status["updated_fields"]:
                            update_jobs[job_id]["updated_fields"][field] = update_jobs[job_id]["updated_fields"].get(field, 0) + 1
                    else:
                        update_jobs[job_id]["skipped"] += 1
                        
                    # Track reasons for updates
                    for reason in update_status.get("reason_for_update", []):
                        update_jobs[job_id]["reasons"][reason] = update_jobs[job_id]["reasons"].get(reason, 0) + 1
                        
                except Exception as e:
                    print(f"Error processing trip {trip_id}: {e}")
                    update_jobs[job_id]["errors"] += 1
                
                update_jobs[job_id]["completed"] += 1
                
        update_jobs[job_id]["status"] = "completed"
        
        # Prepare summary message
        if update_jobs[job_id]["updated"] > 0:
            most_updated_fields = sorted(update_jobs[job_id]["updated_fields"].items(), 
                                         key=lambda x: x[1], reverse=True)[:3]
            update_jobs[job_id]["summary_fields"] = [f"{field} ({count})" for field, count in most_updated_fields]
            
            # Add reasons summary like in process_update_db_async
            most_common_reasons = sorted(update_jobs[job_id]["reasons"].items(), 
                                        key=lambda x: x[1], reverse=True)[:3]
            update_jobs[job_id]["summary_reasons"] = [f"{reason} ({count})" for reason, count in most_common_reasons]
        
    except Exception as e:
        update_jobs[job_id]["status"] = "error"
        update_jobs[job_id]["error_message"] = str(e)

@app.route("/update_progress", methods=["GET"])
def update_progress():
    job_id = request.args.get("job_id")
    if job_id in update_jobs:
        job = update_jobs[job_id]
        total = job.get("total", 0)
        completed = job.get("completed", 0)
        updated = job.get("updated", 0)
        skipped = job.get("skipped", 0)
        percent = (completed / total * 100) if total > 0 else 0
        
        response = {
            "status": job["status"], 
            "total": total, 
            "completed": completed, 
            "percent": percent,
            "updated": updated,
            "skipped": skipped,
            "errors": job.get("errors", 0),
            "created": job.get("created", 0)
        }
        
        # Add summary information if available
        if job["status"] == "completed":
            if "summary_fields" in job:
                response["summary_fields"] = job["summary_fields"]
            if "summary_reasons" in job:
                response["summary_reasons"] = job["summary_reasons"]
            
        return jsonify(response)
    else:
        return jsonify({"error": "Job not found"}), 404

# New endpoint to proxy trip coordinates with proper authentication
@app.route('/trip_coordinates/<int:trip_id>')
def trip_coordinates(trip_id):
    url = f"{BASE_API_URL}/trips/{trip_id}/coordinates"
    # Try to get primary token
    token = fetch_api_token() or API_TOKEN
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    resp = requests.get(url, headers=headers)
    # If unauthorized, try alternative token
    if resp.status_code == 401:
        alt_token = fetch_api_token_alternative()
        if alt_token:
            headers["Authorization"] = f"Bearer {alt_token}"
            resp = requests.get(url, headers=headers)
    try:
        resp.raise_for_status()
        return jsonify(resp.json())
    except Exception as e:
        app.logger.error(f"Error fetching coordinates for trip {trip_id}: {e}")
        return jsonify({"message": "Error fetching coordinates", "error": str(e)}), 500

@app.route("/delete_tag", methods=["POST"])
def delete_tag():
    data = request.get_json()
    tag_name = data.get("name")
    if not tag_name:
        return jsonify(status="error", message="Tag name is required"), 400
    tag = db_session.query(Tag).filter_by(name=tag_name).first()
    if not tag:
        return jsonify(status="error", message="Tag not found"), 404
    # Remove tag from all associated trips
    for trip in list(tag.trips):
        trip.tags.remove(tag)
    db_session.delete(tag)
    db_session.commit()
    return jsonify(status="success", message="Tag deleted successfully")

@app.route("/mixpanel_events", methods=["GET"])
def get_mixpanel_events():
    """
    API endpoint to get Mixpanel events data for the specified date range.
    Query params:
    - start_date: Start date in YYYY-MM-DD format
    - end_date: End date in YYYY-MM-DD format
    """
    from datetime import datetime
    import requests
    import json
    import hashlib
    from flask import request, jsonify
    import os
    
    # Get date range from request parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        return jsonify({"error": "start_date and end_date are required"}), 400
        
    # Create a unique cache key based on the date range
    cache_key = f"mixpanel_events_{start_date}_{end_date}"
    cache_hash = hashlib.md5(cache_key.encode()).hexdigest()
    cache_dir = os.path.join("data", "cache")
    cache_file = os.path.join(cache_dir, f"{cache_hash}.json")
    
    # Create cache directory if it doesn't exist
    if not os.path.exists(cache_dir):
        os.makedirs(cache_dir)
    
    # Check if we have cached data for this date range
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'r') as f:
                cached_data = json.load(f)
                return jsonify(cached_data)
        except Exception as e:
            print(f"Error reading cache file: {e}")
            # If there's an error with the cache, we'll fetch fresh data
    
    # Mixpanel API configuration
    API_SECRET = '725fc2ea9f36a4b3aec9dcbf1b56556d'
    url = "https://data.mixpanel.com/api/2.0/export/"
    
    # Get the event counts
    try:
        # Query parameters for the API request
        params = {
            'from_date': start_date,
            'to_date': end_date
        }
        
        # Headers: specify that we accept JSON
        headers = {
            'Accept': 'application/json'
        }
        
        # Execute the GET request with HTTP Basic Authentication
        response = requests.get(url, auth=(API_SECRET, ''), params=params, headers=headers)
        
        if response.status_code != 200:
            return jsonify({"error": f"Failed to fetch data from Mixpanel: {response.text}"}), 500
        
        # Process each newline-delimited JSON record to get event counts
        event_counts = {}
        for line in response.text.strip().splitlines():
            if line:
                record = json.loads(line)
                event_name = record.get('event')
                if event_name:
                    event_counts[event_name] = event_counts.get(event_name, 0) + 1
        
        # Sort events by counts (descending)
        sorted_events = sorted(
            [{"name": name, "count": count} for name, count in event_counts.items()],
            key=lambda x: x["count"],
            reverse=True
        )
        
        result = {
            "events": sorted_events,
            "start_date": start_date,
            "end_date": end_date,
            "total_count": sum(event_counts.values())
        }
        
        # Cache the result
        try:
            with open(cache_file, 'w') as f:
                json.dump(result, f)
        except Exception as e:
            print(f"Error caching data: {e}")
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({"error": f"Error fetching event data: {str(e)}"}), 500

# ---------------------------
# Impact Analysis Routes
# ---------------------------

@app.route("/impact_analysis")
def impact_analysis():
    """
    Render the impact analysis form page.
    """
    return render_template("impact_analysis.html")

@app.route("/impact_analysis/compare", methods=["POST"])
def impact_analysis_compare():
    """
    Process date ranges and compare metrics between them.
    1. Export data for both date ranges from Mixpanel
    2. Update trip data in the database to ensure it's available 
       (Note: only trips without complete data will be fetched from the API,
       existing trips with complete data will be used from the database)
    3. Process data to get metrics for both periods
    4. Calculate difference and percentage change
    5. Return visualization-ready comparison data
    """
    # Create a unique job ID for this comparison
    job_id = str(uuid.uuid4())
    
    # Initialize progress in global dict
    progress_data[job_id] = {
        "status": "initializing",
        "progress": 0,
        "message": "Starting comparison process...",
        "total_steps": 5,
        "current_step": 0
    }
    
    # Start the process in a background thread to allow progress updates
    thread = threading.Thread(
        target=process_impact_comparison,
        args=(
            job_id,
            request.form.get("base_start_date"),
            request.form.get("base_end_date"),
            request.form.get("comparison_start_date"),
            request.form.get("comparison_end_date")
        )
    )
    thread.daemon = True
    thread.start()
    
    # Return a page that will poll for updates
    return render_template(
        "impact_analysis_progress.html",
        job_id=job_id
    )

def process_impact_comparison(job_id, base_start_date, base_end_date, comparison_start_date, comparison_end_date):
    """
    Background process to handle the impact analysis comparison with progress updates.
    """
    try:
        # Validate date inputs
        if not all([base_start_date, base_end_date, comparison_start_date, comparison_end_date]):
            progress_data[job_id]["status"] = "error"
            progress_data[job_id]["message"] = "All date fields are required."
            return
        
        # Step 1: Export data for both periods from Mixpanel
        update_progress(job_id, 1, "Exporting data from Mixpanel for both time periods...")
        
        base_file, comp_file = export_data_for_comparison(
            base_start_date, base_end_date, 
            comparison_start_date, comparison_end_date
        )
        
        if not base_file or not comp_file:
            progress_data[job_id]["status"] = "error"
            progress_data[job_id]["message"] = "Failed to export data for comparison."
            return
        
        # Step 2: Load data and identify trips that need updates
        update_progress(job_id, 2, "Analyzing which trips need to be updated...")
        
        # Consolidated data update approach
        base_data = load_excel_data(base_file)
        comp_data = load_excel_data(comp_file)
        
        # Get all unique trip IDs from both datasets
        base_trip_ids = [r.get('tripId') for r in base_data if r.get('tripId')]
        comp_trip_ids = [r.get('tripId') for r in comp_data if r.get('tripId')]
        all_trip_ids = list(set(base_trip_ids + comp_trip_ids))
        
        # Check which trips need updating
        session_local = db_session()
        need_updates = False
        try:
            # Get existing trips from database
            existing_trips = {trip.trip_id: trip for trip in session_local.query(Trip).filter(Trip.trip_id.in_(all_trip_ids)).all()}
            
            # Filter out trips that need updates
            missing_data_trips = []
            complete_trips = []
            for trip_id in all_trip_ids:
                if trip_id not in existing_trips:
                    missing_data_trips.append(trip_id)
                elif not _is_trip_data_complete(existing_trips[trip_id]):
                    missing_data_trips.append(trip_id)
                else:
                    complete_trips.append(trip_id)
            
            update_message = f"Found {len(all_trip_ids)} total trips. {len(complete_trips)} already in database, {len(missing_data_trips)} need API updates."
            progress_data[job_id]["details"] = update_message
            app.logger.info(update_message)
            
            need_updates = len(missing_data_trips) > 0
        finally:
            session_local.close()
        
        # Step 3: Update trips that need it
        if need_updates:
            update_progress(job_id, 3, f"Updating {len(missing_data_trips)} trips from API...")
            
            # Create a temporary Excel file with trip IDs for batch update
            consolidated_file = os.path.join("data", "comparison", f"consolidated_trips_{job_id}.xlsx")
            consolidated_df = pd.DataFrame({"tripId": missing_data_trips})
            consolidated_df.to_excel(consolidated_file, index=False)
            
            # Update trips with progress tracking
            # Changed force_update to False since we're already checking which trips need updates
            update_trips_with_progress(consolidated_file, job_id)
            
            # Clean up temp file
            if os.path.exists(consolidated_file):
                os.remove(consolidated_file)
        else:
            update_progress(job_id, 3, "All trips already have complete data in database. No API calls needed.")
        
        # Step 4: Process metrics for both datasets
        update_progress(job_id, 4, "Calculating metrics for both time periods...")
        
        base_metrics = process_data_for_metrics(base_file)
        comparison_metrics = process_data_for_metrics(comp_file)
        
        # Step 5: Calculate comparison results
        update_progress(job_id, 5, "Generating comparison results...")
        
        comparison_results = calculate_comparison_metrics(base_metrics, comparison_metrics)
        
        # Store results in global dict instead of session (which isn't available in background thread)
        progress_data[job_id]["results"] = {
            "comparison_results": comparison_results,
            "comparison_dates": {
                'base_start_date': base_start_date,
                'base_end_date': base_end_date,
                'comparison_start_date': comparison_start_date,
                'comparison_end_date': comparison_end_date
            }
        }
        
        # Mark process as complete
        progress_data[job_id]["status"] = "completed"
        progress_data[job_id]["message"] = "Comparison completed successfully!"
        progress_data[job_id]["progress"] = 100
        
    except Exception as e:
        app.logger.error(f"Error in impact analysis comparison: {str(e)}")
        progress_data[job_id]["status"] = "error"
        progress_data[job_id]["message"] = f"Error processing comparison: {str(e)}"
        traceback.print_exc()

def update_progress(job_id, step, message):
    """Update the progress for a job."""
    if job_id in progress_data:
        progress_data[job_id]["current_step"] = step
        progress_data[job_id]["progress"] = (step / progress_data[job_id]["total_steps"]) * 100
        progress_data[job_id]["message"] = message
        progress_data[job_id]["status"] = "in_progress"
        app.logger.info(f"Step {step}: {message}")

def update_trips_with_progress(excel_file, job_id=None):
    """
    Update trips with progress tracking.
    This function intelligently checks which trips need to be updated:
    - First it checks which trips already have complete data in the database
    - Only trips without complete data will be fetched from the API
    - Uses force_update=False to rely on internal logic in update_trip_db to
      determine if a trip needs updating
    """
    import concurrent.futures
    
    # Load Excel data
    excel_data = load_excel_data(excel_file)
    
    # Get trip IDs from Excel
    trip_ids = [r.get('tripId') for r in excel_data if r.get('tripId')]
    
    if not trip_ids:
        app.logger.warning(f"No trip IDs found in {excel_file}")
        return False
    
    # First check which trips are already in the database with complete data
    session_local = db_session()
    try:
        # Get existing trips from database
        existing_trips = {trip.trip_id: trip for trip in session_local.query(Trip).filter(Trip.trip_id.in_(trip_ids)).all()}
        
        # Filter out trips that already have complete data
        complete_trips = []
        missing_data_trips = []
        for trip_id in trip_ids:
            if trip_id in existing_trips and _is_trip_data_complete(existing_trips[trip_id]):
                complete_trips.append(trip_id)
            else:
                missing_data_trips.append(trip_id)
        
        # Update progress data with information about cache hits
        if job_id and job_id in progress_data:
            db_stats = f"Database: {len(complete_trips)} complete trips found, {len(missing_data_trips)} need API updates"
            progress_data[job_id]["details"] = db_stats
        
        # If all trips are complete, return success immediately
        if not missing_data_trips:
            if job_id and job_id in progress_data:
                progress_data[job_id]["sub_progress"] = f"All {len(trip_ids)} trips already have complete data in database."
            app.logger.info(f"All {len(trip_ids)} trips already have complete data in database.")
            return True
            
        # If we have trips to update, continue only with those
        trip_ids = missing_data_trips
        
    finally:
        session_local.close()
    
    total_count = len(trip_ids)
    success_count = 0
    processed_count = 0
    
    # Define worker function for threaded processing
    def update_single_trip(trip_id):
        nonlocal processed_count, success_count
        
        try:
            # Set force_update to False to rely on internal logic to determine if an update is needed
            db_trip, update_status = update_trip_db(trip_id, force_update=False)
            
            processed_count += 1
            
            # Update progress if job_id provided
            if job_id and job_id in progress_data:
                percent = (processed_count / total_count) * 100
                sub_progress = f"Processed {processed_count}/{total_count} trips from API ({percent:.1f}%)"
                progress_data[job_id]["sub_progress"] = sub_progress
            
            if db_trip and update_status.get("updated_fields"):
                success_count += 1
                return True
            return False
        except Exception as e:
            app.logger.error(f"Failed to update trip {trip_id}: {str(e)}")
            processed_count += 1
            return False
    
    # Use thread pool to update trips in parallel
    max_workers = min(32, (os.cpu_count() or 1) * 4)
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all trips to executor
        future_to_trip = {executor.submit(update_single_trip, trip_id): trip_id for trip_id in trip_ids}
        
        # Process results as they complete
        for future in concurrent.futures.as_completed(future_to_trip):
            pass  # Processing happens in the worker function
    
    # Calculate success rate
    success_rate = success_count / len(missing_data_trips) if missing_data_trips else 1.0
    
    # Log results
    app.logger.info(f"Updated {success_count}/{len(missing_data_trips)} trips from API ({success_rate*100:.1f}%)")
    app.logger.info(f"Total {len(complete_trips)} trips used from database cache")
    
    # Return success if at least 70% of trips were processed successfully or if we had enough cache hits
    return success_rate >= 0.7 or (len(complete_trips) >= 0.7 * (len(complete_trips) + len(missing_data_trips)))

@app.route("/impact_analysis/progress", methods=["GET"])
def impact_comparison_progress():
    """Return progress data for the given job ID."""
    job_id = request.args.get("job_id")
    
    if not job_id or job_id not in progress_data:
        return jsonify({"status": "not_found"})
    
    return jsonify(progress_data[job_id])

@app.route("/impact_analysis/results", methods=["GET"])
def impact_analysis_results():
    """Show the results page after a comparison is complete."""
    job_id = request.args.get("job_id")
    
    if not job_id or job_id not in progress_data or progress_data[job_id]["status"] != "completed":
        return redirect(url_for('impact_analysis'))
    
    # Get results from progress_data
    results_data = progress_data[job_id].get("results", {})
    comparison_results = results_data.get("comparison_results", {})
    dates = results_data.get("comparison_dates", {})
    
    if not comparison_results:
        app.logger.error(f"No comparison results found for job_id: {job_id}")
        return redirect(url_for('impact_analysis'))
    
    # Log data for debugging
    app.logger.debug(f"Chart data for job {job_id}: {comparison_results}")
    
    # Default values
    default_quality_counts = {
        "No Logs Trip": {"base": 1, "comparison": 1, "change": 0, "percent_change": 0},
        "Trip Points Only Exist": {"base": 1, "comparison": 1, "change": 0, "percent_change": 0},
        "Low Quality Trip": {"base": 1, "comparison": 1, "change": 0, "percent_change": 0},
        "Moderate Quality Trip": {"base": 1, "comparison": 1, "change": 0, "percent_change": 0},
        "High Quality Trip": {"base": 1, "comparison": 1, "change": 0, "percent_change": 0},
        "": {"base": 1, "comparison": 1, "change": 0, "percent_change": 0}
    }
    
    default_additional_metrics = {
        "Average Distance Variance": {
            "base": 1.0, "comparison": 1.0, "change": 0.0, "percent_change": 0.0,
            "is_improvement": False, "is_percent": True
        },
        "Accurate Trips %": {
            "base": 80.0, "comparison": 85.0, "change": 5.0, "percent_change": 6.25,
            "is_improvement": True, "is_percent": True
        },
        "App Killed Issue %": {
            "base": 10.0, "comparison": 8.0, "change": -2.0, "percent_change": -20.0,
            "is_improvement": True, "is_percent": True
        },
        "One Log Trips %": {
            "base": 15.0, "comparison": 12.0, "change": -3.0, "percent_change": -20.0,
            "is_improvement": True, "is_percent": True
        },
        "High Quality Trips %": {
            "base": 75.0, "comparison": 80.0, "change": 5.0, "percent_change": 6.67,
            "is_improvement": True, "is_percent": True
        },
        "Low Quality Trips %": {
            "base": 15.0, "comparison": 12.0, "change": -3.0, "percent_change": -20.0,
            "is_improvement": True, "is_percent": True
        }
    }
    
    # Initialize with default empty structures if missing
    if 'quality_counts' not in comparison_results or not comparison_results['quality_counts']:
        comparison_results['quality_counts'] = default_quality_counts.copy()
    if 'avg_manual' not in comparison_results:
        comparison_results['avg_manual'] = {'base': 100.0, 'comparison': 110.0, 'change': 10.0, 'percent_change': 10.0}
    if 'avg_calculated' not in comparison_results:
        comparison_results['avg_calculated'] = {'base': 90.0, 'comparison': 95.0, 'change': 5.0, 'percent_change': 5.56}
    if 'additional_metrics' not in comparison_results or not comparison_results['additional_metrics']:
        comparison_results['additional_metrics'] = default_additional_metrics.copy()
    
    # Ensure all required metrics are present
    for metric, values in default_additional_metrics.items():
        if metric not in comparison_results['additional_metrics']:
            comparison_results['additional_metrics'][metric] = values.copy()
    
    # Ensure all quality categories are present
    for category, values in default_quality_counts.items():
        if category not in comparison_results['quality_counts']:
            comparison_results['quality_counts'][category] = values.copy()
    
    # Sanitize all data structures to ensure they're JSON-serializable
    
    # For quality_counts
    sanitized_quality_counts = {}
    for quality, values in comparison_results['quality_counts'].items():
        sanitized_values = {}
        for key, value in values.items():
            try:
                # Convert numpy types or ensure float/int as appropriate
                if hasattr(value, 'item'):
                    sanitized_values[key] = value.item()
                elif key in ['base', 'comparison', 'change']:
                    # Ensure at least 1 for chart display
                    sanitized_values[key] = max(1, int(float(value)) if value is not None else 1)
                else:
                    sanitized_values[key] = float(value) if value is not None else 0.0
            except (ValueError, TypeError):
                # Default values if conversion fails
                if key in ['base', 'comparison', 'change']:
                    sanitized_values[key] = 1
                else:
                    sanitized_values[key] = 0.0
        sanitized_quality_counts[quality] = sanitized_values
    comparison_results['quality_counts'] = sanitized_quality_counts
    
    # For avg_manual and avg_calculated
    for metric_key in ['avg_manual', 'avg_calculated']:
        sanitized_metric = {}
        for key, value in comparison_results[metric_key].items():
            try:
                if hasattr(value, 'item'):
                    sanitized_metric[key] = value.item()
                else:
                    sanitized_metric[key] = float(value) if value is not None else 0.0
            except (ValueError, TypeError):
                sanitized_metric[key] = 0.0
        comparison_results[metric_key] = sanitized_metric
    
    # For additional_metrics
    sanitized_additional_metrics = {}
    for metric, values in comparison_results['additional_metrics'].items():
        sanitized_values = {}
        for key, value in values.items():
            if key in ['is_improvement', 'is_percent']:
                # Boolean values should remain boolean
                sanitized_values[key] = bool(value)
            else:
                try:
                    if hasattr(value, 'item'):
                        sanitized_values[key] = value.item()
                    else:
                        sanitized_values[key] = float(value) if value is not None else 0.0
                except (ValueError, TypeError):
                    sanitized_values[key] = 0.0
        sanitized_additional_metrics[metric] = sanitized_values
    comparison_results['additional_metrics'] = sanitized_additional_metrics
    
    # Ensure all dates are strings
    sanitized_dates = {}
    for key, value in dates.items():
        sanitized_dates[key] = str(value)
    
    return render_template(
        "impact_analysis.html",
        comparison_results=comparison_results,
        base_start_date=sanitized_dates.get('base_start_date', ''),
        base_end_date=sanitized_dates.get('base_end_date', ''),
        comparison_start_date=sanitized_dates.get('comparison_start_date', ''),
        comparison_end_date=sanitized_dates.get('comparison_end_date', '')
    )

def _is_trip_data_complete(trip):
    """
    Check if a trip record has all the necessary data for analysis.
    
    Args:
        trip: Trip database object
        
    Returns:
        bool: True if the trip has all the needed data, False otherwise
    """
    # Check if trip is None
    if trip is None:
        return False
        
    # Check for essential fields
    required_numeric_fields = [
        'manual_distance',
        'calculated_distance',
        'short_segments_count',
        'medium_segments_count',
        'long_segments_count',
        'short_segments_distance',
        'medium_segments_distance',
        'long_segments_distance',
        'coordinate_count'
    ]
    
    required_string_fields = [
        'route_quality',
        'expected_trip_quality',
        'device_type',
        'carrier'
    ]
    
    # Check numeric fields - they should exist and be convertible to float
    for field in required_numeric_fields:
        if not hasattr(trip, field) or getattr(trip, field) is None:
            return False
        try:
            value = getattr(trip, field)
            if value == "":
                return False
            float(value)  # Try to convert to float
        except (ValueError, TypeError):
            return False
    
    # Check string fields - they should exist and not be empty
    for field in required_string_fields:
        if not hasattr(trip, field) or getattr(trip, field) is None:
            return False
        if str(getattr(trip, field)).strip() == "":
            return False
            
    # Check boolean fields
    if not hasattr(trip, 'lack_of_accuracy'):
        return False
    
    return True

def process_data_for_metrics(excel_file):
    """
    Process Excel data to extract key metrics for comparison.
    Returns a dictionary with all metrics.
    """
    from collections import Counter, defaultdict
    
    # Load excel data
    excel_data = load_excel_data(excel_file)
    
    # Get trip IDs from Excel
    excel_trip_ids = [r.get('tripId') for r in excel_data if r.get('tripId')]
    
    # Initialize metrics with default values
    # Use at least 1 for each quality count to ensure chart data is populated
    metrics = {
        "quality_counts": {
            "No Logs Trip": 1,
            "Trip Points Only Exist": 1,
            "Low Quality Trip": 1,
            "Moderate Quality Trip": 1,
            "High Quality Trip": 1,
            "": 1  # for empty quality
        },
        "total_manual": 0.0,
        "total_calculated": 0.0,
        "count_manual": 0,
        "count_calculated": 0,
        "consistent": 0,
        "inconsistent": 0,
        "variance_sum": 0.0,
        "variance_count": 0,
        "accurate_count": 0,  # <25% variance
        "app_killed_count": 0,
        "one_log_count": 0,
        "total_short_dist": 0.0,
        "total_medium_dist": 0.0,
        "total_long_dist": 0.0,
        "total_trip_count": 0,
        # Additional metrics for automatic insights
        "total_coordinate_count": 0,
        "total_trip_duration": 0.0,
        "trip_duration_count": 0,
        "avg_coordinate_count": 0,
        "avg_trip_duration": 0.0
    }
    
    # If no trip IDs found, return default metrics
    if not excel_trip_ids:
        app.logger.warning(f"No trip IDs found in {excel_file}")
        return metrics
    
    # Connect to DB
    session_local = db_session()
    
    # Query trips in Excel
    try:
        trips_db = session_local.query(Trip).filter(Trip.trip_id.in_(excel_trip_ids)).all()
        
        # If no trips found in database, return default metrics
        if not trips_db:
            app.logger.warning(f"No trips found in database for IDs in {excel_file}")
            return metrics
        
        # Reset quality counts to 0 since we have real data
        metrics["quality_counts"] = {
            "No Logs Trip": 0,
            "Trip Points Only Exist": 0,
            "Low Quality Trip": 0,
            "Moderate Quality Trip": 0,
            "High Quality Trip": 0,
            "": 0  # for empty quality
        }
        
        # Filter out trips with calculated_distance > 2000 km
        filtered_trips = []
        for trip in trips_db:
            try:
                cd = float(trip.calculated_distance) if trip.calculated_distance is not None else 0
                if cd <= 2000:
                    filtered_trips.append(trip)
            except (ValueError, TypeError):
                continue
        
        metrics["total_trip_count"] = len(filtered_trips)
        
        # Process trip metrics
        for trip in filtered_trips:
            # Count quality categories
            eq_quality = (trip.expected_trip_quality or "").strip()
            if eq_quality in metrics["quality_counts"]:
                metrics["quality_counts"][eq_quality] += 1
            else:
                metrics["quality_counts"][""] += 1
            
            # Distance metrics
            try:
                md = float(trip.manual_distance) if trip.manual_distance is not None else 0
                cd = float(trip.calculated_distance) if trip.calculated_distance is not None else 0
                metrics["total_manual"] += md
                metrics["total_calculated"] += cd
                metrics["count_manual"] += 1
                metrics["count_calculated"] += 1
                
                if md > 0:
                    variance = abs(cd - md) / md * 100
                    metrics["variance_sum"] += variance
                    metrics["variance_count"] += 1
                    if variance < 25.0:
                        metrics["accurate_count"] += 1
                
                if md > 0 and abs(cd - md) / md <= 0.2:
                    metrics["consistent"] += 1
                else:
                    metrics["inconsistent"] += 1
            except:
                pass
            
            # Segment distances
            if trip.short_segments_distance:
                try:
                    metrics["total_short_dist"] += float(trip.short_segments_distance)
                except (ValueError, TypeError):
                    pass
            if trip.medium_segments_distance:
                try:
                    metrics["total_medium_dist"] += float(trip.medium_segments_distance)
                except (ValueError, TypeError):
                    pass
            if trip.long_segments_distance:
                try:
                    metrics["total_long_dist"] += float(trip.long_segments_distance)
                except (ValueError, TypeError):
                    pass
            
            # Coordinate count for automatic insights
            if trip.coordinate_count:
                try:
                    count = int(trip.coordinate_count)
                    metrics["total_coordinate_count"] += count
                    if count == 1:
                        metrics["one_log_count"] += 1
                except (ValueError, TypeError):
                    pass
            
            # Trip duration for automatic insights
            if trip.trip_time:
                try:
                    raw_tt = float(trip.trip_time)
                    # Convert possible seconds→hours if over 72
                    if raw_tt > 72:
                        raw_tt /= 3600.0
                    # Skip if >720 hours or negative
                    if raw_tt >= 0 and raw_tt <= 720:
                        metrics["total_trip_duration"] += raw_tt
                        metrics["trip_duration_count"] += 1
                except (ValueError, TypeError):
                    pass
            
            # App killed issue
            if trip.lack_of_accuracy:
                metrics["app_killed_count"] += 1
        
        # If we have real data but no quality counts, restore default values
        if sum(metrics["quality_counts"].values()) == 0:
            metrics["quality_counts"] = {
                "No Logs Trip": 1,
                "Trip Points Only Exist": 1, 
                "Low Quality Trip": 1,
                "Moderate Quality Trip": 1,
                "High Quality Trip": 1,
                "": 1
            }
        
        # Calculate averages for automatic insights metrics
        if metrics["total_trip_count"] > 0:
            metrics["avg_coordinate_count"] = metrics["total_coordinate_count"] / metrics["total_trip_count"]
        
        if metrics["trip_duration_count"] > 0:
            metrics["avg_trip_duration"] = metrics["total_trip_duration"] / metrics["trip_duration_count"]
        
    except Exception as e:
        app.logger.error(f"Error processing metrics from {excel_file}: {str(e)}")
    finally:
        session_local.close()
    
    return metrics

def calculate_comparison_metrics(base_metrics, comparison_metrics):
    """
    Calculate differences and percentage changes between base and comparison metrics.
    Returns formatted data for rendering in template.
    """
    results = {
        "quality_counts": {},
        "avg_manual": {},
        "avg_calculated": {},
        "additional_metrics": {},
        "automatic_insights": {}  # Add automatic insights section
    }
    
    # Ensure default metrics are present
    default_metrics = [
        "Average Distance Variance",
        "Accurate Trips %",
        "App Killed Issue %",
        "One Log Trips %",
        "High Quality Trips %",
        "Low Quality Trips %"
    ]
    
    # Ensure all quality categories are present in both metrics
    quality_categories = ["No Logs Trip", "Trip Points Only Exist", "Low Quality Trip", 
                         "Moderate Quality Trip", "High Quality Trip", ""]
    
    for quality in quality_categories:
        if quality not in base_metrics["quality_counts"]:
            base_metrics["quality_counts"][quality] = 0
        if quality not in comparison_metrics["quality_counts"]:
            comparison_metrics["quality_counts"][quality] = 0
    
    # Quality counts comparison
    for quality in quality_categories:
        base_count = base_metrics["quality_counts"].get(quality, 0)
        comp_count = comparison_metrics["quality_counts"].get(quality, 0)
        change = comp_count - base_count
        # Avoid division by zero
        percent_change = (change / max(base_count, 1) * 100) if base_count != 0 else 0
        
        results["quality_counts"][quality] = {
            "base": int(base_count),
            "comparison": int(comp_count),
            "change": int(change),
            "percent_change": float(percent_change)
        }
    
    # Distance averages
    base_avg_manual = base_metrics["total_manual"] / max(base_metrics["count_manual"], 1) if base_metrics["count_manual"] > 0 else 0
    comp_avg_manual = comparison_metrics["total_manual"] / max(comparison_metrics["count_manual"], 1) if comparison_metrics["count_manual"] > 0 else 0
    manual_change = comp_avg_manual - base_avg_manual
    # Avoid division by zero
    manual_percent_change = (manual_change / max(base_avg_manual, 0.0001) * 100) if base_avg_manual > 0 else 0
    
    results["avg_manual"] = {
        "base": float(base_avg_manual),
        "comparison": float(comp_avg_manual),
        "change": float(manual_change),
        "percent_change": float(manual_percent_change)
    }
    
    base_avg_calculated = base_metrics["total_calculated"] / max(base_metrics["count_calculated"], 1) if base_metrics["count_calculated"] > 0 else 0
    comp_avg_calculated = comparison_metrics["total_calculated"] / max(comparison_metrics["count_calculated"], 1) if comparison_metrics["count_calculated"] > 0 else 0
    calculated_change = comp_avg_calculated - base_avg_calculated
    # Avoid division by zero
    calculated_percent_change = (calculated_change / max(base_avg_calculated, 0.0001) * 100) if base_avg_calculated > 0 else 0
    
    results["avg_calculated"] = {
        "base": float(base_avg_calculated),
        "comparison": float(comp_avg_calculated),
        "change": float(calculated_change),
        "percent_change": float(calculated_percent_change)
    }
    
    # Additional metrics comparison
    
    # 1. Average distance variance (lower is better)
    base_avg_variance = base_metrics["variance_sum"] / max(base_metrics["variance_count"], 1) if base_metrics["variance_count"] > 0 else 0
    comp_avg_variance = comparison_metrics["variance_sum"] / max(comparison_metrics["variance_count"], 1) if comparison_metrics["variance_count"] > 0 else 0
    variance_change = comp_avg_variance - base_avg_variance
    # Avoid division by zero
    variance_percent_change = (variance_change / max(base_avg_variance, 0.0001) * 100) if base_avg_variance > 0 else 0
    
    results["additional_metrics"]["Average Distance Variance"] = {
        "base": float(base_avg_variance),
        "comparison": float(comp_avg_variance),
        "change": float(variance_change),
        "percent_change": float(variance_percent_change),
        "is_improvement": variance_change < 0,  # Lower variance is better
        "is_percent": True
    }
    
    # 2. Accurate trips percentage (higher is better)
    base_accurate_pct = (base_metrics["accurate_count"] / max(base_metrics["total_trip_count"], 1) * 100) if base_metrics["total_trip_count"] > 0 else 0
    comp_accurate_pct = (comparison_metrics["accurate_count"] / max(comparison_metrics["total_trip_count"], 1) * 100) if comparison_metrics["total_trip_count"] > 0 else 0
    accurate_change = comp_accurate_pct - base_accurate_pct
    # Avoid division by zero
    accurate_percent_change = (accurate_change / max(base_accurate_pct, 0.0001) * 100) if base_accurate_pct > 0 else 0
    
    results["additional_metrics"]["Accurate Trips %"] = {
        "base": float(base_accurate_pct),
        "comparison": float(comp_accurate_pct),
        "change": float(accurate_change),
        "percent_change": float(accurate_percent_change),
        "is_improvement": accurate_change > 0,  # Higher accuracy percentage is better
        "is_percent": True
    }
    
    # 3. App killed percentage (lower is better)
    base_app_killed_pct = (base_metrics["app_killed_count"] / max(base_metrics["total_trip_count"], 1) * 100) if base_metrics["total_trip_count"] > 0 else 0
    comp_app_killed_pct = (comparison_metrics["app_killed_count"] / max(comparison_metrics["total_trip_count"], 1) * 100) if comparison_metrics["total_trip_count"] > 0 else 0
    app_killed_change = comp_app_killed_pct - base_app_killed_pct
    # Avoid division by zero
    app_killed_percent_change = (app_killed_change / max(base_app_killed_pct, 0.0001) * 100) if base_app_killed_pct > 0 else 0
    
    results["additional_metrics"]["App Killed Issue %"] = {
        "base": float(base_app_killed_pct),
        "comparison": float(comp_app_killed_pct),
        "change": float(app_killed_change),
        "percent_change": float(app_killed_percent_change),
        "is_improvement": app_killed_change < 0,  # Lower app killed percentage is better
        "is_percent": True
    }
    
    # 4. One log percentage (lower is better)
    base_one_log_pct = (base_metrics["one_log_count"] / max(base_metrics["total_trip_count"], 1) * 100) if base_metrics["total_trip_count"] > 0 else 0
    comp_one_log_pct = (comparison_metrics["one_log_count"] / max(comparison_metrics["total_trip_count"], 1) * 100) if comparison_metrics["total_trip_count"] > 0 else 0
    one_log_change = comp_one_log_pct - base_one_log_pct
    # Avoid division by zero
    one_log_percent_change = (one_log_change / max(base_one_log_pct, 0.0001) * 100) if base_one_log_pct > 0 else 0
    
    results["additional_metrics"]["One Log Trips %"] = {
        "base": float(base_one_log_pct),
        "comparison": float(comp_one_log_pct),
        "change": float(one_log_change),
        "percent_change": float(one_log_percent_change),
        "is_improvement": one_log_change < 0,  # Lower one log percentage is better
        "is_percent": True
    }
    
    # 5. High quality percentage (higher is better)
    base_high_quality_pct = (base_metrics["quality_counts"].get("High Quality Trip", 0) / max(base_metrics["total_trip_count"], 1) * 100) if base_metrics["total_trip_count"] > 0 else 0
    comp_high_quality_pct = (comparison_metrics["quality_counts"].get("High Quality Trip", 0) / max(comparison_metrics["total_trip_count"], 1) * 100) if comparison_metrics["total_trip_count"] > 0 else 0
    high_quality_change = comp_high_quality_pct - base_high_quality_pct
    # Avoid division by zero
    high_quality_percent_change = (high_quality_change / max(base_high_quality_pct, 0.0001) * 100) if base_high_quality_pct > 0 else 0
    
    results["additional_metrics"]["High Quality Trips %"] = {
        "base": float(base_high_quality_pct),
        "comparison": float(comp_high_quality_pct),
        "change": float(high_quality_change),
        "percent_change": float(high_quality_percent_change),
        "is_improvement": high_quality_change > 0,  # Higher high quality percentage is better
        "is_percent": True
    }
    
    # 6. Low quality percentage (lower is better)
    base_low_quality_pct = (base_metrics["quality_counts"].get("Low Quality Trip", 0) / max(base_metrics["total_trip_count"], 1) * 100) if base_metrics["total_trip_count"] > 0 else 0
    comp_low_quality_pct = (comparison_metrics["quality_counts"].get("Low Quality Trip", 0) / max(comparison_metrics["total_trip_count"], 1) * 100) if comparison_metrics["total_trip_count"] > 0 else 0
    low_quality_change = comp_low_quality_pct - base_low_quality_pct
    # Avoid division by zero
    low_quality_percent_change = (low_quality_change / max(base_low_quality_pct, 0.0001) * 100) if base_low_quality_pct > 0 else 0
    
    results["additional_metrics"]["Low Quality Trips %"] = {
        "base": float(base_low_quality_pct),
        "comparison": float(comp_low_quality_pct),
        "change": float(low_quality_change),
        "percent_change": float(low_quality_percent_change),
        "is_improvement": low_quality_change < 0,  # Lower low quality percentage is better
        "is_percent": True
    }
    
    # Verify all required metrics are present
    for metric in default_metrics:
        if metric not in results["additional_metrics"]:
            results["additional_metrics"][metric] = {
                "base": 0.0,
                "comparison": 0.0,
                "change": 0.0,
                "percent_change": 0.0,
                "is_improvement": False,
                "is_percent": True
            }
    
    # Add automatic insights metrics
    # Average Logs Count
    base_logs_avg = base_metrics.get("avg_coordinate_count", 0)
    comp_logs_avg = comparison_metrics.get("avg_coordinate_count", 0)
    logs_change = comp_logs_avg - base_logs_avg
    logs_percent_change = (logs_change / max(base_logs_avg, 0.0001) * 100) if base_logs_avg > 0 else 0
    
    results["automatic_insights"]["Average Logs Count"] = {
        "base": float(base_logs_avg),
        "comparison": float(comp_logs_avg),
        "change": float(logs_change),
        "percent_change": float(logs_percent_change),
        "is_improvement": logs_change > 0,  # Higher logs count is better
        "is_percent": False
    }
    
    # Average Trip Duration
    base_duration_avg = base_metrics.get("avg_trip_duration", 0)
    comp_duration_avg = comparison_metrics.get("avg_trip_duration", 0)
    duration_change = comp_duration_avg - base_duration_avg
    duration_percent_change = (duration_change / max(base_duration_avg, 0.0001) * 100) if base_duration_avg > 0 else 0
    
    results["automatic_insights"]["Average Trip Duration"] = {
        "base": float(base_duration_avg),
        "comparison": float(comp_duration_avg),
        "change": float(duration_change),
        "percent_change": float(duration_percent_change),
        "is_improvement": False,  # Duration by itself is not better/worse
        "is_percent": False
    }
    
    # Distance Segment Distribution 
    base_short_dist_pct = (base_metrics.get("total_short_dist", 0) / max(base_metrics["total_calculated"], 0.0001) * 100) if base_metrics["total_calculated"] > 0 else 0
    comp_short_dist_pct = (comparison_metrics.get("total_short_dist", 0) / max(comparison_metrics["total_calculated"], 0.0001) * 100) if comparison_metrics["total_calculated"] > 0 else 0
    short_dist_change = comp_short_dist_pct - base_short_dist_pct
    short_dist_percent_change = (short_dist_change / max(base_short_dist_pct, 0.0001) * 100) if base_short_dist_pct > 0 else 0
    
    results["automatic_insights"]["Short Segments %"] = {
        "base": float(base_short_dist_pct),
        "comparison": float(comp_short_dist_pct),
        "change": float(short_dist_change),
        "percent_change": float(short_dist_percent_change),
        "is_improvement": short_dist_change < 0,  # Lower short segments is better
        "is_percent": True
    }
    
    # Medium Segments
    base_medium_dist_pct = (base_metrics.get("total_medium_dist", 0) / max(base_metrics["total_calculated"], 0.0001) * 100) if base_metrics["total_calculated"] > 0 else 0
    comp_medium_dist_pct = (comparison_metrics.get("total_medium_dist", 0) / max(comparison_metrics["total_calculated"], 0.0001) * 100) if comparison_metrics["total_calculated"] > 0 else 0
    medium_dist_change = comp_medium_dist_pct - base_medium_dist_pct
    medium_dist_percent_change = (medium_dist_change / max(base_medium_dist_pct, 0.0001) * 100) if base_medium_dist_pct > 0 else 0
    
    results["automatic_insights"]["Medium Segments %"] = {
        "base": float(base_medium_dist_pct),
        "comparison": float(comp_medium_dist_pct),
        "change": float(medium_dist_change),
        "percent_change": float(medium_dist_percent_change),
        "is_improvement": medium_dist_change >= 0,  # Higher medium segments is better
        "is_percent": True
    }
    
    # Long Segments
    base_long_dist_pct = (base_metrics.get("total_long_dist", 0) / max(base_metrics["total_calculated"], 0.0001) * 100) if base_metrics["total_calculated"] > 0 else 0
    comp_long_dist_pct = (comparison_metrics.get("total_long_dist", 0) / max(comparison_metrics["total_calculated"], 0.0001) * 100) if comparison_metrics["total_calculated"] > 0 else 0
    long_dist_change = comp_long_dist_pct - base_long_dist_pct
    long_dist_percent_change = (long_dist_change / max(base_long_dist_pct, 0.0001) * 100) if base_long_dist_pct > 0 else 0
    
    results["automatic_insights"]["Long Segments %"] = {
        "base": float(base_long_dist_pct),
        "comparison": float(comp_long_dist_pct),
        "change": float(long_dist_change),
        "percent_change": float(long_dist_percent_change),
        "is_improvement": long_dist_change >= 0,  # Higher long segments is better
        "is_percent": True
    }
    
    return results

@app.route("/download_driver_logs/<int:trip_id>", methods=["POST"])
def download_driver_logs(trip_id):
    """
    Download driver logs for a specific trip, analyze them for issues, and store the results.
    
    The function will fetch logs from the API, look for common issues such as:
    - MQTT connection issues
    - Network connectivity problems
    - App crashes
    - Memory pressure indicators
    - Location tracking failures
    
    Returns JSON with analysis results and tags.
    """
    try:
        # Retrieve trip from database
        session_local = db_session()
        trip = session_local.query(Trip).filter(Trip.trip_id == trip_id).first()
        
        if not trip:
            return jsonify({"status": "error", "message": f"Trip {trip_id} not found"}), 404

        # Get driver ID from associated excel data
        excel_path = os.path.join("data", "data.xlsx")
        excel_data = load_excel_data(excel_path)
        trip_data = next((r for r in excel_data if r.get("tripId") == trip_id), None)
        
        if not trip_data:
            return jsonify({"status": "error", "message": f"Trip {trip_id} not found in excel data"}), 404
        
        driver_id = trip_data.get("UserId")
        trip_date = trip_data.get("time")
        
        if not driver_id:
            return jsonify({"status": "error", "message": "Driver ID not found for this trip"}), 404
        
        if not trip_date:
            return jsonify({"status": "error", "message": "Trip date not found"}), 404
        
        # Convert trip_date to datetime if it's a string
        if isinstance(trip_date, str):
            try:
                trip_date = datetime.strptime(trip_date, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return jsonify({"status": "error", "message": "Invalid trip date format"}), 400
        
        # Make the API request
        download_token = "eyJhbGciOiJub25lIn0.eyJpZCI6MTgsIm5hbWUiOiJUZXN0IERyaXZlciIsInBob25lX251bWJlciI6IisyMDEwMDA2Mjk5OTgiLCJwaG90byI6eyJ1cmwiOm51bGx9LCJkcml2ZXJfbGljZW5zZSI6eyJ1cmwiOm51bGx9LCJjcmVhdGVkX2F0IjoiMjAxOS0wMy0xMyAwMDoyMjozMiArMDIwMCIsInVwZGF0ZWRfYXQiOiIyMDE5LTAzLTEzIDAwOjIyOjMyICswMjAwIiwibmF0aW9uYWxfaWQiOiIxMjM0NSIsImVtYWlsIjoicHJvZEBwcm9kLmNvbSIsImdjbV9kZXZpY2VfdG9rZW4iOm51bGx9."
        headers = {
            "Authorization": f"Bearer {download_token}",
            "Content-Type": "application/json"
        }
        
        # API endpoint for driver logs
        api_url = f"https://app.illa.blue/api/v3/driver/driver_app_logs?filter[driver_id]={driver_id}&all_pages=true"
        
        response = requests.get(api_url, headers=headers)
        
        if response.status_code != 200:
            # Try with alternative token
            alt_token = fetch_api_token_alternative()
            if alt_token:
                headers["Authorization"] = f"Bearer {alt_token}"
                response = requests.get(api_url, headers=headers)
                
                if response.status_code != 200:
                    return jsonify({
                        "status": "error",
                        "message": f"Failed to fetch logs: {response.status_code}"
                    }), response.status_code
            else:
                return jsonify({
                    "status": "error",
                    "message": f"Failed to fetch logs: {response.status_code}"
                }), response.status_code
        
        # Process the response
        logs_data = response.json()
        
        # Check if logs are in the 'data' field instead of 'logs' field
        log_items = logs_data.get("logs", [])
        if not log_items and "data" in logs_data:
            log_items = logs_data.get("data", [])
            
        if not log_items:
            return jsonify({
                "status": "error",
                "message": "No log files found for this driver. The driver may not have submitted any logs, or there might be an issue with the driver ID."
            }), 404
        
        # Define a function to parse various datetime formats from the API
        def parse_datetime(date_str):
            formats_to_try = [
                "%Y-%m-%dT%H:%M:%S%z",     # ISO 8601 with timezone
                "%Y-%m-%dT%H:%M:%S.%f%z",  # ISO 8601 with ms and timezone
                "%Y-%m-%dT%H:%M:%SZ",      # ISO 8601 with Z
                "%Y-%m-%dT%H:%M:%S.%fZ",   # ISO 8601 with ms and Z
                "%Y-%m-%dT%H:%M:%S",       # ISO 8601 without timezone
                "%Y-%m-%dT%H:%M:%S.%f",    # ISO 8601 with ms, without timezone
                "%Y-%m-%d %H:%M:%S",       # Simple datetime
                "%Y-%m-%d %H:%M:%S%z"      # Simple datetime with timezone
            ]
            
            for fmt in formats_to_try:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    # Remove timezone info to make it offset-naive
                    if dt.tzinfo is not None:
                        dt = dt.replace(tzinfo=None)
                    return dt
                except ValueError:
                    continue
            
            # If we reach here, none of the formats matched
            raise ValueError(f"Could not parse datetime string: {date_str}")
        
        # Create a list to store logs with their parsed dates
        logs_with_dates = []
        
        for log in log_items:
            # Extract the created date based on response structure
            created_date_str = None
            
            # Check if the log has 'attributes' field (JSON:API format)
            if isinstance(log, dict) and "attributes" in log:
                attributes = log.get("attributes", {})
                if "createdAt" in attributes:
                    created_date_str = attributes.get("createdAt")
                elif "created_at" in attributes:
                    created_date_str = attributes.get("created_at")
            # Direct access for simple JSON format
            elif isinstance(log, dict):
                if "createdAt" in log:
                    created_date_str = log.get("createdAt")
                elif "created_at" in log:
                    created_date_str = log.get("created_at")
            
            if not created_date_str:
                continue
            
            try:
                created_date = parse_datetime(created_date_str)
                logs_with_dates.append((log, created_date))
            except ValueError:
                continue
        
        if not logs_with_dates:
            return jsonify({
                "status": "error",
                "message": "No logs with valid dates found for this driver."
            }), 404
        
        # Sort logs by date
        logs_with_dates.sort(key=lambda x: x[1])
        
        # Define a time window to look for logs (12 hours before and after the trip)
        time_window_start = trip_date - timedelta(hours=12)
        time_window_end = trip_date + timedelta(hours=12)
        
        # Find logs within the time window
        logs_in_window = [
            (log, log_date) for log, log_date in logs_with_dates 
            if time_window_start <= log_date <= time_window_end
        ]
        
        # If no logs in window, try a larger window (24 hours)
        if not logs_in_window:
            time_window_start = trip_date - timedelta(hours=24)
            time_window_end = trip_date + timedelta(hours=24)
            logs_in_window = [
                (log, log_date) for log, log_date in logs_with_dates 
                if time_window_start <= log_date <= time_window_end
            ]
        
        # If still no logs in the expanded window, try an even larger window (48 hours)
        if not logs_in_window:
            time_window_start = trip_date - timedelta(hours=48)
            time_window_end = trip_date + timedelta(hours=48)
            logs_in_window = [
                (log, log_date) for log, log_date in logs_with_dates 
                if time_window_start <= log_date <= time_window_end
            ]
        
        # If there are logs in the window, use the closest one to the trip date
        if logs_in_window:
            closest_log = min(logs_in_window, key=lambda x: abs((x[1] - trip_date).total_seconds()))[0]
        else:
            # If no logs in any window, use the closest one by date
            closest_log = min(logs_with_dates, key=lambda x: abs((x[1] - trip_date).total_seconds()))[0]
        
        # Get the log file URL based on the response structure
        log_file_url = None
        
        if "attributes" in closest_log and "logFileUrl" in closest_log["attributes"]:
            log_file_url = closest_log["attributes"]["logFileUrl"]
        elif "logFileUrl" in closest_log:
            log_file_url = closest_log["logFileUrl"]
            
        if not log_file_url:
            return jsonify({
                "status": "error",
                "message": "Log file URL not found in the API response. The log file might be missing or corrupted."
            }), 404
        
        log_response = requests.get(log_file_url)
        if log_response.status_code != 200:
            return jsonify({
                "status": "error",
                "message": f"Failed to download log file: {log_response.status_code}"
            }), log_response.status_code
        
        # Save log file
        # Get filename based on response structure
        log_filename = None
        if "attributes" in closest_log and "filename" in closest_log["attributes"]:
            log_filename = closest_log["attributes"]["filename"]
        elif "filename" in closest_log:
            log_filename = closest_log["filename"]
            
        if not log_filename:
            log_filename = f"log_{trip_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.txt"
            
        log_path = os.path.join("data", log_filename)
        
        with open(log_path, "wb") as f:
            f.write(log_response.content)
        
        # Analyze the log file
        log_content = log_response.content
        try:
            # Try to decode as UTF-8 first
            log_content = log_response.content.decode('utf-8')
        except UnicodeDecodeError:
            # If it's not UTF-8, try to decompress if it's a gzip file
            if log_filename.endswith('.gz'):
                import gzip
                import io
                try:
                    with gzip.GzipFile(fileobj=io.BytesIO(log_response.content)) as f:
                        log_content = f.read().decode('utf-8', errors='replace')
                except Exception:
                    # If decompression fails, use raw content with errors replaced
                    log_content = log_response.content.decode('utf-8', errors='replace')
            else:
                # Not a gzip file, use raw content with errors replaced
                log_content = log_response.content.decode('utf-8', errors='replace')
                
        analysis_results = analyze_log_file(log_content, trip_id)
        
        # Save analysis results to trip record
        if analysis_results.get("tags"):
            # Convert tags to Tag objects if they don't exist
            for tag_name in analysis_results["tags"]:
                tag = session_local.query(Tag).filter(Tag.name == tag_name).first()
                if not tag:
                    tag = Tag(name=tag_name)
                    session_local.add(tag)
                    session_local.flush()
                
                # Add tag to trip if not already present
                if tag not in trip.tags:
                    trip.tags.append(tag)
        
        session_local.commit()
        
        return jsonify({
            "status": "success",
            "message": "Log file downloaded and analyzed successfully",
            "filename": log_filename,
            "analysis": analysis_results
        })
    
    except Exception as e:
        import traceback
        return jsonify({
            "status": "error",
            "message": f"An error occurred: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500
    finally:
        session_local.close()

def analyze_log_file(log_content, trip_id):
    """
    Analyze the log file content for common issues.
    
    Args:
        log_content: The text content of the log file
        trip_id: The ID of the trip
        
    Returns:
        Dictionary with analysis results including tags and time periods
    """
    lines = log_content.split('\n')
    analysis = {
        "tags": [],
        "total_lines": len(lines),
        "time_without_logs": 0,  # in seconds
        "first_timestamp": None,
        "last_timestamp": None,
        "mqtt_connection_issues": 0,
        "network_connectivity_issues": 0,
        "location_tracking_issues": 0,
        "memory_pressure_indicators": 0,
        "app_crashes": 0,
        "server_errors": 0,
        "battery_optimizations": 0,
        "background_time": 0,  # in seconds
        "foreground_time": 0,  # in seconds
        "app_sessions": 0,
        "task_removals": 0,  # times app was removed from recents
        "gps_toggles": 0,  # times GPS was turned on/off
        "network_toggles": 0,  # times network connectivity changed
        "background_transitions": 0,  # times app went to background
        "foreground_transitions": 0,  # times app came to foreground
        "location_sync_attempts": 0,
        "location_sync_failures": 0,
        "trip_events": [],  # important events during trip in chronological order
    }
    
    # Track application state
    app_state = {
        "is_in_foreground": False,
        "last_state_change": None,
        "current_network_state": None,
        "is_tracking_active": False,
        "last_timestamp": None
    }
    
    # Regular expressions for extracting timestamps and specific log patterns
    timestamp_pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})'
    timestamps = []
    
    # Process each line
    for line in lines:
        # Extract timestamp if available
        timestamp_match = re.search(timestamp_pattern, line)
        if timestamp_match:
            timestamp = timestamp_match.group(1)
            timestamps.append(timestamp)
            
            # Update first and last timestamp
            if not analysis["first_timestamp"]:
                analysis["first_timestamp"] = timestamp
            analysis["last_timestamp"] = timestamp
            
            # Update app state timestamp
            if app_state["last_timestamp"] and timestamp != app_state["last_timestamp"]:
                app_state["last_timestamp"] = timestamp
            elif not app_state["last_timestamp"]:
                app_state["last_timestamp"] = timestamp
        
        # Check for MQTT connection issues
        if "MqttException" in line or ("MQTT" in line and "failure" in line):
            analysis["mqtt_connection_issues"] += 1
            analysis["trip_events"].append({
                "time": timestamp if timestamp_match else None,
                "event": "MQTT Connection Issue",
                "details": line.strip()
            })
        
        # Check for network connectivity issues
        if "UnknownHostException" in line or "SocketTimeoutException" in line:
            analysis["network_connectivity_issues"] += 1
            analysis["trip_events"].append({
                "time": timestamp if timestamp_match else None,
                "event": "Network Connectivity Issue",
                "details": line.strip()
            })
        
        # Track network state changes
        if "NetworkConnectivityReceiver" in line and "Network status changed" in line:
            analysis["network_toggles"] += 1
            new_state = "Connected" if "Connected" in line else "Disconnected"
            
            if app_state["current_network_state"] != new_state:
                app_state["current_network_state"] = new_state
                analysis["trip_events"].append({
                    "time": timestamp if timestamp_match else None,
                    "event": f"Network Changed to {new_state}",
                    "details": line.strip()
                })
        
        # Check for location tracking issues
        if "Location tracking" in line and ("failed" in line or "error" in line):
            analysis["location_tracking_issues"] += 1
            analysis["trip_events"].append({
                "time": timestamp if timestamp_match else None,
                "event": "Location Tracking Issue",
                "details": line.strip()
            })
        
        # Track location sync attempts and failures
        if "LocationSyncWorker" in line and "Syncing locations" in line:
            analysis["location_sync_attempts"] += 1
            # Try to extract number of locations being synced
            locations_count_match = re.search(r'Syncing \[(\d+)\] locations', line)
            if locations_count_match:
                locations_count = int(locations_count_match.group(1))
                analysis["trip_events"].append({
                    "time": timestamp if timestamp_match else None,
                    "event": f"Location Sync Attempt",
                    "details": f"Attempted to sync {locations_count} locations"
                })
            
        if "LocationSyncWorker" in line and ("failed" in line or "Error" in line):
            analysis["location_sync_failures"] += 1
            analysis["trip_events"].append({
                "time": timestamp if timestamp_match else None,
                "event": "Location Sync Failure",
                "details": line.strip()
            })
        
        # Check for memory pressure
        if "onTrimMemory" in line or "memory pressure" in line:
            analysis["memory_pressure_indicators"] += 1
            # Extract memory trim level if available
            trim_level_match = re.search(r'onTrimMemory called with Level= (\w+)', line)
            if trim_level_match and trim_level_match.group(1) == "TRIM_MEMORY_COMPLETE":
                analysis["trip_events"].append({
                    "time": timestamp if timestamp_match else None,
                    "event": "Severe Memory Pressure",
                    "details": "System requested complete memory trimming"
                })
        
        # Check for app crashes
        if "FATAL EXCEPTION" in line or "crash" in line or "ANR" in line:
            analysis["app_crashes"] += 1
            analysis["trip_events"].append({
                "time": timestamp if timestamp_match else None,
                "event": "App Crash",
                "details": line.strip()
            })
        
        # Check for server errors
        if "HTTP 5" in line or "server error" in line:
            analysis["server_errors"] += 1
            analysis["trip_events"].append({
                "time": timestamp if timestamp_match else None,
                "event": "Server Error",
                "details": line.strip()
            })
        
        # Track app foreground/background transitions
        if "BackgroundDetector" in line:
            if "app is in Foreground : true" in line or "ActivityResumed, app is in Foreground : true" in line:
                if not app_state["is_in_foreground"]:
                    analysis["foreground_transitions"] += 1
                    app_state["is_in_foreground"] = True
                    app_state["last_state_change"] = timestamp if timestamp_match else None
                    analysis["trip_events"].append({
                        "time": timestamp if timestamp_match else None,
                        "event": "App To Foreground",
                        "details": "Application moved to foreground"
                    })
            elif "app is in inBackground : true" in line or "Activity-Stopped, app is in inBackground : true" in line:
                if app_state["is_in_foreground"]:
                    analysis["background_transitions"] += 1
                    app_state["is_in_foreground"] = False
                    app_state["last_state_change"] = timestamp if timestamp_match else None
                    analysis["trip_events"].append({
                        "time": timestamp if timestamp_match else None,
                        "event": "App To Background",
                        "details": "Application moved to background"
                    })
        
        # Track app session starts
        if "illa" in line and "Logging Started" in line:
            analysis["app_sessions"] += 1
            analysis["trip_events"].append({
                "time": timestamp if timestamp_match else None,
                "event": "App Session Started",
                "details": "New application session began"
            })
        
        # Track onTaskRemoved events (user swipes app away from recents)
        if "onTaskRemoved" in line:
            analysis["task_removals"] += 1
            analysis["trip_events"].append({
                "time": timestamp if timestamp_match else None,
                "event": "App Removed From Recents",
                "details": "User removed app from recent apps list"
            })
        
        # Track trip start/end events
        if "TrackingService" in line and "tracking state -> [Started]" in line:
            app_state["is_tracking_active"] = True
            analysis["trip_events"].append({
                "time": timestamp if timestamp_match else None,
                "event": "Trip Tracking Started",
                "details": line.strip()
            })
            
        if "TrackingService" in line and "tracking state -> [Stopped]" in line:
            app_state["is_tracking_active"] = False
            analysis["trip_events"].append({
                "time": timestamp if timestamp_match else None,
                "event": "Trip Tracking Stopped",
                "details": line.strip()
            })
            
        # Track GPS state changes
        if "LocationManagerProvider" in line:
            if "Location updates requested" in line:
                analysis["gps_toggles"] += 1
                analysis["trip_events"].append({
                    "time": timestamp if timestamp_match else None,
                    "event": "GPS Tracking Enabled",
                    "details": "Location updates were requested"
                })
            elif "Location updates removed" in line:
                analysis["gps_toggles"] += 1
                analysis["trip_events"].append({
                    "time": timestamp if timestamp_match else None,
                    "event": "GPS Tracking Disabled",
                    "details": "Location updates were stopped"
                })
        
        # Check for battery optimization messages
        if "battery" in line.lower() and ("optimization" in line.lower() or "doze" in line.lower()):
            analysis["battery_optimizations"] += 1
            analysis["trip_events"].append({
                "time": timestamp if timestamp_match else None,
                "event": "Battery Optimization",
                "details": line.strip()
            })
    
    # Calculate time without logs if we have at least 2 timestamps
    if len(timestamps) >= 2:
        # Convert to datetime objects
        datetime_timestamps = []
        for ts in timestamps:
            try:
                # Extract date from timestamp
                dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
                datetime_timestamps.append(dt)
            except ValueError:
                continue
        
        # Sort timestamps
        datetime_timestamps.sort()
        
        # Check for gaps and calculate total time
        for i in range(1, len(datetime_timestamps)):
            time_diff = (datetime_timestamps[i] - datetime_timestamps[i-1]).total_seconds()
            if time_diff > 300:  # Gap of more than 5 minutes
                analysis["time_without_logs"] += time_diff
                analysis["trip_events"].append({
                    "time": timestamps[i-1],
                    "event": "Log Gap",
                    "details": f"No logs for {time_diff:.1f} seconds until {timestamps[i]}"
                })
    
    # Calculate total trip duration if we have a first and last timestamp
    if analysis["first_timestamp"] and analysis["last_timestamp"]:
        try:
            first_dt = datetime.strptime(analysis["first_timestamp"], "%Y-%m-%d %H:%M:%S")
            last_dt = datetime.strptime(analysis["last_timestamp"], "%Y-%m-%d %H:%M:%S")
            analysis["total_duration"] = (last_dt - first_dt).total_seconds()
        except ValueError:
            analysis["total_duration"] = 0
    
    # Determine tags based on issue counts
    if analysis["mqtt_connection_issues"] > 50:
        analysis["tags"].append("MQTT Connection Issues")
        
    if analysis["network_connectivity_issues"] > 20:
        analysis["tags"].append("Network Connectivity Issues")
        
    if analysis["location_tracking_issues"] > 100:
        analysis["tags"].append("Location Tracking Issues")
        
    if analysis["memory_pressure_indicators"] > 15:
        analysis["tags"].append("Memory Pressure")
        
    if analysis["app_crashes"] > 0:
        analysis["tags"].append("App Crashes")
        
    if analysis["server_errors"] > 0:
        analysis["tags"].append("Server Communication Issues")
    
    if analysis["task_removals"] > 0:
        analysis["tags"].append("App Removed From Recents")
    
    if analysis["background_transitions"] > 10:
        analysis["tags"].append("Frequent Background Transitions")
    
    if analysis["app_sessions"] > 5:
        analysis["tags"].append("Multiple App Sessions")
    
    if analysis["location_sync_failures"] > 0:
        analysis["tags"].append("Location Sync Failures")
    
    if analysis["time_without_logs"] > 1200:  # More than 5 minutes
        analysis["tags"].append("Significant Log Gaps")
    
    if analysis["battery_optimizations"] > 0:
        analysis["tags"].append("Battery Optimization Detected")
    
    # Detect trip end status
    if "trip ended" in log_content.lower() or "trip terminated" in log_content.lower() or "tracking state -> [Stopped]" in log_content:
        analysis["tags"].append("Normal Trip Termination")
    
    # Detect if logs show kill by OS
    if "process killed" in log_content.lower() or "killed by system" in log_content.lower():
        analysis["tags"].append("Killed by OS")
    
    # Detect if the app was in the background
    if "app is in inBackground : true" in log_content or "Activity-Stopped, app is in inBackground : true" in log_content:
        analysis["tags"].append("App Background Transitions")
    
    # Detect if locations were synchronized
    if "locations synced" in log_content.lower() or "successfully synced locations" in log_content.lower():
        analysis["tags"].append("Successful Location Sync")
    
    # Sort trip events chronologically
    analysis["trip_events"].sort(key=lambda x: x["time"] if x["time"] else "")
    
    return analysis

@app.route("/update_all_trips_tags", methods=["POST"])
def update_all_trips_tags():
    """
    Updates all trip tags by analyzing the log files for each trip.
    
    This function will:
    1. Get all trips from the Excel file
    2. For each trip, download the log file if available
    3. Analyze the log file to identify issues
    4. Apply tags to the trip based on the analysis
    
    Returns JSON with update statistics.
    """
    job_id = f"update_tags_{int(time.time())}"
    update_jobs[job_id] = {
        "status": "in_progress",
        "total": 0,
        "completed": 0,
        "updated": 0,
        "skipped": 0,
        "errors": 0,
        "percent": 0,
        "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    # Start processing in a background thread
    thread = Thread(target=process_update_all_trips_tags, args=(job_id,))
    thread.daemon = True
    thread.start()
    
    return jsonify({
        "status": "started",
        "job_id": job_id,
        "message": "Update tags process started."
    })

def process_update_all_trips_tags(job_id):
    """
    Background process to analyze trip logs and update tags for trips in the Excel file.
    Uses concurrent.futures to process trips in parallel.
    """
    try:
        # Get excel data
        excel_path = os.path.join("data", "data.xlsx")
        excel_data = load_excel_data(excel_path)
        update_jobs[job_id]["total"] = len(excel_data)
        
        # Use ThreadPoolExecutor to process trips in parallel
        with concurrent.futures.ThreadPoolExecutor(max_workers=40) as executor:
            # Submit all trips for processing
            future_to_trip = {
                executor.submit(process_single_trip_tag_update, trip_data, job_id): trip_data.get("tripId")
                for trip_data in excel_data if trip_data.get("tripId")
            }
            
            # Process completed tasks
            for future in concurrent.futures.as_completed(future_to_trip):
                trip_id = future_to_trip[future]
                try:
                    future.result()  # Get any exceptions
                except Exception as e:
                    app.logger.error(f"Error processing trip {trip_id}: {str(e)}")
                    update_jobs[job_id]["errors"] += 1
                finally:
                    # Update progress
                    update_jobs[job_id]["percent"] = min(100, (update_jobs[job_id]["completed"] * 100) / max(1, update_jobs[job_id]["total"]))
        
        update_jobs[job_id]["status"] = "completed"
        
    except Exception as e:
        update_jobs[job_id]["status"] = "error"
        update_jobs[job_id]["error_message"] = str(e)

def process_single_trip_tag_update(trip_data, job_id):
    """
    Process a single trip for tag update.
    """
    session_local = None
    try:
        session_local = db_session()
        trip_id = trip_data.get("tripId")
        
        # Check if trip exists in the database
        trip = session_local.query(Trip).filter(Trip.trip_id == trip_id).first()
        if not trip:
            app.logger.warning(f"Trip {trip_id} not found in database, skipping tag analysis")
            update_jobs[job_id]["skipped"] += 1
            update_jobs[job_id]["completed"] += 1
            return
        
        driver_id = trip_data.get("UserId")
        trip_date = trip_data.get("time")
        
        if not driver_id or not trip_date:
            app.logger.warning(f"Missing driver ID or trip date for trip {trip_id}, skipping tag analysis")
            update_jobs[job_id]["skipped"] += 1
            update_jobs[job_id]["completed"] += 1
            return
        
        # Convert trip_date to datetime if it's a string
        if isinstance(trip_date, str):
            try:
                trip_date = datetime.strptime(trip_date, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                app.logger.error(f"Invalid trip date format for trip {trip_id}")
                update_jobs[job_id]["errors"] += 1
                update_jobs[job_id]["completed"] += 1
                return
        
        # Make the API request for driver logs
        download_token = "eyJhbGciOiJub25lIn0.eyJpZCI6MTgsIm5hbWUiOiJUZXN0IERyaXZlciIsInBob25lX251bWJlciI6IisyMDEwMDA2Mjk5OTgiLCJwaG90byI6eyJ1cmwiOm51bGx9LCJkcml2ZXJfbGljZW5zZSI6eyJ1cmwiOm51bGx9LCJjcmVhdGVkX2F0IjoiMjAxOS0wMy0xMyAwMDoyMjozMiArMDIwMCIsInVwZGF0ZWRfYXQiOiIyMDE5LTAzLTEzIDAwOjIyOjMyICswMjAwIiwibmF0aW9uYWxfaWQiOiIxMjM0NSIsImVtYWlsIjoicHJvZEBwcm9kLmNvbSIsImdjbV9kZXZpY2VfdG9rZW4iOm51bGx9."
        headers = {
            "Authorization": f"Bearer {download_token}",
            "Content-Type": "application/json"
        }
        
        # API endpoint for driver logs
        api_url = f"https://app.illa.blue/api/v3/driver/driver_app_logs?filter[driver_id]={driver_id}&all_pages=true"
        
        response = requests.get(api_url, headers=headers)
        
        if response.status_code != 200:
            # Try with alternative token
            alt_token = fetch_api_token_alternative()
            if alt_token:
                headers["Authorization"] = f"Bearer {alt_token}"
                response = requests.get(api_url, headers=headers)
                
                if response.status_code != 200:
                    app.logger.error(f"Failed to fetch logs for trip {trip_id}: {response.status_code}")
                    update_jobs[job_id]["errors"] += 1
                    update_jobs[job_id]["completed"] += 1
                    return
            else:
                app.logger.error(f"Failed to fetch logs for trip {trip_id}: {response.status_code}")
                update_jobs[job_id]["errors"] += 1
                update_jobs[job_id]["completed"] += 1
                return
        
        # Process the response
        logs_data = response.json()
        
        # Check if logs are in the 'data' field instead of 'logs' field
        log_items = logs_data.get("logs", [])
        if not log_items and "data" in logs_data:
            log_items = logs_data.get("data", [])
            
        if not log_items:
            app.logger.warning(f"No log files found for trip {trip_id}")
            update_jobs[job_id]["skipped"] += 1
            update_jobs[job_id]["completed"] += 1
            return
        
        # Parse datetime function
        def parse_datetime(date_str):
            formats_to_try = [
                "%Y-%m-%dT%H:%M:%S%z",     # ISO 8601 with timezone
                "%Y-%m-%dT%H:%M:%S.%f%z",  # ISO 8601 with ms and timezone
                "%Y-%m-%dT%H:%M:%SZ",      # ISO 8601 with Z
                "%Y-%m-%dT%H:%M:%S.%fZ",   # ISO 8601 with ms and Z
                "%Y-%m-%dT%H:%M:%S",       # ISO 8601 without timezone
                "%Y-%m-%dT%H:%M:%S.%f",    # ISO 8601 with ms, without timezone
                "%Y-%m-%d %H:%M:%S",       # Simple datetime
                "%Y-%m-%d %H:%M:%S%z"      # Simple datetime with timezone
            ]
            
            for fmt in formats_to_try:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    # Remove timezone info to make it offset-naive
                    if dt.tzinfo is not None:
                        dt = dt.replace(tzinfo=None)
                    return dt
                except ValueError:
                    continue
            
            # If we reach here, none of the formats matched
            raise ValueError(f"Could not parse datetime string: {date_str}")
        
        # Create a list to store logs with their parsed dates
        logs_with_dates = []
        
        for log in log_items:
            # Extract the created date based on response structure
            created_date_str = None
            
            # Check if the log has 'attributes' field (JSON:API format)
            if isinstance(log, dict) and "attributes" in log:
                attributes = log.get("attributes", {})
                if "createdAt" in attributes:
                    created_date_str = attributes.get("createdAt")
                elif "created_at" in attributes:
                    created_date_str = attributes.get("created_at")
            # Direct access for simple JSON format
            elif isinstance(log, dict):
                if "createdAt" in log:
                    created_date_str = log.get("createdAt")
                elif "created_at" in log:
                    created_date_str = log.get("created_at")
            
            if not created_date_str:
                continue
            
            try:
                created_date = parse_datetime(created_date_str)
                logs_with_dates.append((log, created_date))
            except ValueError:
                continue
        
        if not logs_with_dates:
            app.logger.warning(f"No logs with valid dates found for trip {trip_id}")
            update_jobs[job_id]["skipped"] += 1
            update_jobs[job_id]["completed"] += 1
            return
        
        # Sort logs by date
        logs_with_dates.sort(key=lambda x: x[1])
        
        # Define a time window to look for logs (12 hours before and after the trip)
        time_window_start = trip_date - timedelta(hours=12)
        time_window_end = trip_date + timedelta(hours=12)
        
        # Find logs within the time window
        logs_in_window = [
            (log, log_date) for log, log_date in logs_with_dates 
            if time_window_start <= log_date <= time_window_end
        ]
        
        # If no logs in window, try a larger window (24 hours)
        if not logs_in_window:
            time_window_start = trip_date - timedelta(hours=24)
            time_window_end = trip_date + timedelta(hours=24)
            logs_in_window = [
                (log, log_date) for log, log_date in logs_with_dates 
                if time_window_start <= log_date <= time_window_end
            ]
        
        # If still no logs in the expanded window, try an even larger window (48 hours)
        if not logs_in_window:
            time_window_start = trip_date - timedelta(hours=48)
            time_window_end = trip_date + timedelta(hours=48)
            logs_in_window = [
                (log, log_date) for log, log_date in logs_with_dates 
                if time_window_start <= log_date <= time_window_end
            ]
        
        # If there are logs in the window, use the closest one to the trip date
        if logs_in_window:
            closest_log = min(logs_in_window, key=lambda x: abs((x[1] - trip_date).total_seconds()))[0]
        else:
            # If no logs in any window, use the closest one by date
            closest_log = min(logs_with_dates, key=lambda x: abs((x[1] - trip_date).total_seconds()))[0]
        
        # Get the log file URL based on the response structure
        log_file_url = None
        
        if "attributes" in closest_log and "logFileUrl" in closest_log["attributes"]:
            log_file_url = closest_log["attributes"]["logFileUrl"]
        elif "logFileUrl" in closest_log:
            log_file_url = closest_log["logFileUrl"]
            
        if not log_file_url:
            app.logger.warning(f"Log file URL not found for trip {trip_id}")
            update_jobs[job_id]["skipped"] += 1
            update_jobs[job_id]["completed"] += 1
            return
        
        log_response = requests.get(log_file_url)
        if log_response.status_code != 200:
            app.logger.error(f"Failed to download log file for trip {trip_id}: {log_response.status_code}")
            update_jobs[job_id]["errors"] += 1
            update_jobs[job_id]["completed"] += 1
            return
        
        # Get filename based on response structure
        log_filename = None
        if "attributes" in closest_log and "filename" in closest_log["attributes"]:
            log_filename = closest_log["attributes"]["filename"]
        elif "filename" in closest_log:
            log_filename = closest_log["filename"]
            
        if not log_filename:
            log_filename = f"log_{trip_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.txt"
            
        log_path = os.path.join("data", log_filename)
        
        with open(log_path, "wb") as f:
            f.write(log_response.content)
        
        # Analyze the log file
        log_content = log_response.content
        try:
            # Try to decode as UTF-8 first
            log_content = log_response.content.decode('utf-8')
        except UnicodeDecodeError:
            # If it's not UTF-8, try to decompress if it's a gzip file
            if log_filename.endswith('.gz'):
                import gzip
                import io
                try:
                    with gzip.GzipFile(fileobj=io.BytesIO(log_response.content)) as f:
                        log_content = f.read().decode('utf-8', errors='replace')
                except Exception:
                    # If decompression fails, use raw content with errors replaced
                    log_content = log_response.content.decode('utf-8', errors='replace')
            else:
                # Not a gzip file, use raw content with errors replaced
                log_content = log_response.content.decode('utf-8', errors='replace')
                
        analysis_results = analyze_log_file(log_content, trip_id)
        
        # Save analysis results to trip record
        if analysis_results.get("tags"):
            # Clear existing tags
            trip.tags = []
            
            # Convert tags to Tag objects
            for tag_name in analysis_results["tags"]:
                tag = session_local.query(Tag).filter(Tag.name == tag_name).first()
                if not tag:
                    tag = Tag(name=tag_name)
                    session_local.add(tag)
                    session_local.flush()
                
                # Add tag to trip
                trip.tags.append(tag)
            
            session_local.commit()
            update_jobs[job_id]["updated"] += 1
        else:
            update_jobs[job_id]["skipped"] += 1
        
    except Exception as e:
        app.logger.error(f"Error processing trip {trip_data.get('tripId')}: {str(e)}")
        update_jobs[job_id]["errors"] += 1
    finally:
        update_jobs[job_id]["completed"] += 1
        if session_local:
            session_local.close()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
